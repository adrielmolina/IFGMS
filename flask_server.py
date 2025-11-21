from flask import Flask, request, render_template, redirect, url_for, flash, session, jsonify, send_file, send_from_directory, abort
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from livereload import Server
from py_scripts import db_conn, tools, models
from py_scripts.db_conn import SessionLocal
from datetime import date, datetime
from sqlalchemy import create_engine, text, func, extract, case, case, and_, or_
import os
import pandas as pd
import openpyxl
from functools import wraps
from io import BytesIO
from flask import send_file
import requests


server = Flask(__name__)
server.jinja_env.auto_reload = True
server.secret_key = os.urandom(24)
#reCaptcha
SITE_KEY = os.getenv("SITE_KEY")
if not SITE_KEY:
    print("reCaptcha site key not set.")
    raise ValueError("SITE_KEY environment variable not set.")

SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    print("reCaptcha secret key not set.")
    raise ValueError("SECRET_KEY environment variable not set.")


# CACHE CONTROL FOR STATIC FILES
cache_bypass = True

if cache_bypass or os.getenv("FLASK_ENV") == "production":
    server.config['SEND_FILE_MAX_AGE_DEFAULT'] = 31536000
else:
    server.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0


login_manager = LoginManager()
login_manager.init_app(server)
login_manager.login_view = 'landing_page'
login_manager.login_message = {
    "title": "Login Required",
    "text": "You must log in to view this page.",
    "redirect_url": "/"
}
# TODO remove all flash template and use another dialog box
# TODO try toastr for the dialogs
login_manager.login_message_category = "warning"

# TODO put @login_required on all template and api routes
#! TODO redirect user to dashboard if already logged in and trying to access login page

# for closing the session after requests
@server.teardown_appcontext
def cleanup(exception=None):
    db_conn.shutdown_session()

# f0r the login 
@login_manager.user_loader
def load_user(user_id):
    db_session = db_conn.SessionLocal()
    return db_session.query(db_conn.models.Accounts).get(int(user_id))


def roles_required(*roles):
    '''  Decorator to restrict access to users with a specific role.'''
    def wrapper(fn):
        @wraps(fn)
        def decorated_view(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.user_level not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return decorated_view
    return wrapper


# ?TODO LIST-------------------------------
# TODO add the summary per month to the module list or wherever

# TODO disable this before deployment
@server.before_request
def auto_login():
    ''' Auto-login for development purposes. Remove or disable in production. '''
    auto_login_enabled = False if os.getenv("FLASK_ENV") == "production" else True
    
    db_session = db_conn.SessionLocal()
    if server.config.get("DEBUG_BYPASS_LOGIN", auto_login_enabled):
        if not current_user.is_authenticated:
            test_user = db_session.query(db_conn.models.Accounts).filter_by(username="adriel").first()
            login_user(test_user)


#? -------------------- LOGIN / LOGOUT -------------------- ?#

# TODO add the usernamee and pass to the route address
@server.route('/login', methods=['POST'])
def login():
    username = request.form.get("username")
    password = request.form.get("password")
    token = request.form.get("g-recaptcha-response")

    # 2️⃣ Verify captcha first
    if not token:
        flash({
            "title": "Login Error!",
            "text": "Please complete the CAPTCHA.",
            "redirect_url": url_for('landing_page')
        }, "error")
        return render_template('index.html')

    resp = requests.post(
        "https://www.google.com/recaptcha/api/siteverify",
        data={"secret": SECRET_KEY, "response": token}
    ).json()

    if not resp.get("success"):
        flash({
            "title": "Login Error!",
            "text": "CAPTCHA verification failed. Try again.",
            "redirect_url": url_for('landing_page')
        }, "error")
        return render_template('index.html')
    
    user = db_conn.sign_in(username, password)

    if user:
        # Refresh the user object with a new session for Flask-Login
        db_session = db_conn.SessionLocal()
        try:
            refreshed_user = db_session.query(db_conn.models.Accounts).get(user.account_id)
            
            print(f'🔍 DEBUG: User found: {refreshed_user.username}, Status: {refreshed_user.acct_status}')
            if refreshed_user.acct_status == 'approved':
                login_user(refreshed_user)
                db_conn.POST_action_log(refreshed_user.username, refreshed_user.user_level, 'Login Attempt', 'Success', refreshed_user.account_id)
                return redirect(url_for('dashboard'))

            elif refreshed_user.acct_status == 'pending':
                flash({
                    "title": "Login Error!",
                    "text": "Account not approved yet. Contact admin.",
                    "redirect_url": url_for('landing_page')
                },"error")
                db_conn.POST_action_log(username, None, 'Login Attempt', 'Fail. Account status pending', None)
                return render_template('index.html')
        finally:
            db_session.close()
    else:
        flash({
            "title": "Login Error!",
            "text": "Wrong username or password. Try again.",
            "redirect_url": url_for('landing_page')
        }, "error")
        db_conn.POST_action_log(username, None, 'Login Attempt', 'Fail. Wrong username/password', None)
        return render_template('index.html')


@server.route('/logout')
@login_required
def logout():
    db_conn.POST_action_log(current_user.username, current_user.user_level, 'Logout', 'User logged out', current_user.account_id)
    logout_user()
    session.clear()
    return redirect(url_for("landing_page"))



#? -------------------- END -------------------- ?#

#? -------------------- NAVIGATIONS -------------------- ?#

@server.route('/')
def landing_page():
    # redirect to dashboard if already logged in
    if os.getenv("FLASK_ENV") == "production" or os.getenv("FLASK_ENV") == "staging":
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        
    
    if os.getenv("env") == "production":
        env = 'Live'
    elif os.getenv("env") == "staging":
        env = 'Staging'
    else:
        env = 'Development'
    
    return render_template('index.html', env=env, site_key=SITE_KEY)


@server.route('/create_account')
def create_acc():
    return render_template('create_account.html')


@server.route('/forgot_password')
def forgot_password():
    return render_template('forgot_password.html')


@server.route('/membership_register')
def membership_register():
    return render_template('membership_register.html', site_key=SITE_KEY)


@server.route('/profile_settings')
@login_required
def profile_settings():    
    return render_template('profile_settings.html')


# TODO UNUSED. REMOVE IF NOT NEEDED
@server.route('/profile')
@login_required
def profile():
    return render_template('profile.html')


@server.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')


@server.route('/members')
@login_required
def members_page():
    user_location = current_user.office_location if current_user and current_user.office_location else 'Chapter'
    user_level = current_user.user_level if current_user else 'staff'
    is_chapter_user = user_location == 'Chapter'
    
    print(f"DEBUG: User location: {user_location}, User level: {user_level}, Is Chapter: {is_chapter_user}")
    
    return render_template('members.html', 
                         user_location=user_location, 
                         user_level=user_level,
                         is_chapter_user=is_chapter_user)

@server.route('/declaration')
@login_required
@roles_required('admin', 'superadmin')
def declaration_page():
    
    active_dispatch = db_conn.get_current_active_dispatch()
    if active_dispatch:
        dispatch_contents = db_conn.get_current_dispatch_contents(active_dispatch.dispatch_id)
        
        # Get unique categories and locations FROM THE CURRENT DISPATCH CONTENTS only
        unique_categories = set()
        unique_locations = set()
        
        for row in dispatch_contents:
            if hasattr(row, 'maab_category') and row.maab_category:
                unique_categories.add(row.maab_category)
            if hasattr(row, 'location_particular') and row.location_particular:
                unique_locations.add(row.location_particular)
        
        # Convert to sorted lists
        categories_list = sorted(list(unique_categories))
        locations_list = sorted(list(unique_locations))
        
        if dispatch_contents:      
            print('✅ Current dispatch:', active_dispatch.dispatch_id)
            print('✅ Dispatch contents count:', len(dispatch_contents))
            print('✅ Categories in dispatch:', categories_list)
            print('✅ Locations in dispatch:', locations_list)
            return render_template('declaration.html', 
                                active_dispatch=active_dispatch, 
                                dispatch_contents=dispatch_contents,
                                categories=categories_list,
                                locations=locations_list)
        else:
            # if empty or error
            print('⚠️ No dispatch contents found')
            return render_template('declaration.html', 
                                active_dispatch=active_dispatch, 
                                dispatch_contents=[],
                                categories=[],
                                locations=[])
    else:
        print('❌ No active dispatch found')
        return render_template('declaration.html', 
                            active_dispatch=False,
                            categories=[],
                            locations=[])

@server.route('/inventory')
@login_required
def inventory():
    # TODO move to a fetch at page load
    # Fetch inventory entries from the database
    inventory_data = db_conn.get_inventory_entries()  # Ensure this returns data correctly
    
    # Print or log the data to verify it's being passed correctly
    print(inventory_data)
    
    # Pass data to the template
    return render_template('inventory.html', inventory_data=inventory_data)


@server.route('/claims')
@login_required
@roles_required('admin', 'superadmin')
def claims():
    if current_user.office_location != 'Chapter':
        abort(403)
    return render_template('claims.html')


@server.route('/reports')
@login_required
@roles_required('admin', 'superadmin')
def bud_v_exp():
    return render_template('reports.html')


@server.route('/audit_trails')
@login_required
@roles_required('admin', 'superadmin')
def audit_trails():
    # TODO move to a fetch at page load
    logs = db_conn.GET_audit_logs()
    
    return render_template('audit_trails.html', logs=logs)


@server.route('/accounts')
@login_required
@roles_required('admin', 'superadmin')
def show_user_accounts():
    return render_template('accounts.html')


@server.route('/settings')
@login_required
def settings():
    has_profile_pic = current_user.profile_pic is not None
    return render_template('settings.html', has_profile_pic=has_profile_pic)


#? -------------------- END -------------------- ?#

#? -------------------- API ROUTES -------------------- ?#
@server.route('/api/dashboard/sales_data', methods=['GET'])
@login_required
def get_sales_data():
    """Get current sales data for dashboard - FIXED CALCULATION"""
    try:
        db_session = SessionLocal()
        
        # Get current year
        current_year = datetime.now().year
        
        # Get all paid entries for current year with their categories
        paid_entries = db_session.query(
            models.Entries.maab_category,
            func.count(models.Entries.entry_id).label('count')
        ).filter(
            extract('year', models.Entries.OR_date) == current_year,
            models.Entries.paid == True
        ).group_by(models.Entries.maab_category).all()
        
        # Price mapping for each membership type
        price_map = {
            'Classic': 60,
            'Bronze': 150,
            'Silver': 300,
            'Gold': 500,
            'Platinum': 1000,
            'Safe Card': 1000,  # Enhanced Platinum
            'Senior': 300,
            'Senior+': 350
        }
        
        # Calculate TOTAL ACTUAL SALES (sum of price × count for all categories)
        total_actual_sales = 0
        for category, count in paid_entries:
            price = price_map.get(category, 0)
            total_actual_sales += count * price
        
        # Get target data for current year
        target_data = db_session.query(models.Report_TvA).filter(
            models.Report_TvA.year == current_year
        ).first()
        
        # Calculate TOTAL TARGET SALES (sum of target × price for all categories)
        total_target_sales = 0
        if target_data:
            total_target_sales = (
                (target_data.classic or 0) * price_map['Classic'] +
                (target_data.bronze or 0) * price_map['Bronze'] +
                (target_data.silver or 0) * price_map['Silver'] +
                (target_data.gold or 0) * price_map['Gold'] +
                (target_data.platinum or 0) * price_map['Platinum'] +
                (target_data.safe_card or 0) * price_map['Safe Card'] +
                (target_data.senior or 0) * price_map['Senior'] +
                (target_data.senior_plus or 0) * price_map['Senior+']
            )
        else:
            # Default target if not set - calculate based on reasonable defaults
            default_targets = {
                'Classic': 30000,    # 30000 × 60 = 1,800,000
                'Bronze': 75000,     # 75000 × 150 = 11,250,000  
                'Silver': 15000,     # 15000 × 300 = 4,500,000
                'Gold': 25000,       # 25000 × 500 = 12,500,000
                'Platinum': 50000,   # 50000 × 1000 = 50,000,000
                'Safe Card': 60000,  # 60000 × 1000 = 60,000,000
                'Senior': 15000,     # 15000 × 300 = 4,500,000
                'Senior+': 17500     # 17500 × 350 = 6,125,000
            }
            total_target_sales = sum(
                default_targets[cat] * price_map[cat] 
                for cat in default_targets.keys()
            )
        
        print(f"💰 SALES CALCULATION DEBUG:")
        print(f"   Current Year: {current_year}")
        print(f"   Paid Entries: {paid_entries}")
        print(f"   Total Actual Sales: ₱{total_actual_sales:,}")
        print(f"   Total Target Sales: ₱{total_target_sales:,}")
        
        return jsonify({
            'success': True,
            'current_total': total_actual_sales,  # Total sales amount in pesos
            'target_total': total_target_sales,   # Total target amount in pesos
            'year': current_year
        })
        
    except Exception as e:
        print(f"Error getting sales data: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'current_total': 0,
            'target_total': 1000000  # Fallback target
        }), 500
    finally:
        db_session.close()

@server.route('/api/dashboard/sales_performance', methods=['GET'])
@login_required
def get_sales_performance():
    """Get sales performance data by membership type for chart"""
    try:
        db_session = SessionLocal()
        current_year = datetime.now().year
        
        # Get actual sales count per category for current year
        actual_sales = db_session.query(
            models.Entries.maab_category,
            func.count(models.Entries.entry_id).label('actual_count')
        ).filter(
            extract('year', models.Entries.OR_date) == current_year,
            models.Entries.paid == True
        ).group_by(models.Entries.maab_category).all()
        
        # Get targets for current year
        target_data = db_session.query(models.Report_TvA).filter(
            models.Report_TvA.year == current_year
        ).first()
        
        # Map category names to match your chart
        category_mapping = {
            'Classic': 'Classic',
            'Bronze': 'Premier Bronze', 
            'Silver': 'Premier Silver',
            'Gold': 'Premier Gold',
            'Platinum': 'Premier Platinum',
            'Safe Card': 'Safe Card',  # or 'Safe Card' based on your preference
            'Senior': 'Senior',
            'Senior+': 'Senior Plus'
        }
        
        # Prepare chart data
        chart_data = []
        all_categories = [
            'Classic', 'Premier Bronze', 'Premier Silver', 'Premier Gold', 
            'Premier Platinum', 'Safe Card', 'Senior', 'Senior Plus'
        ]
        
        for category in all_categories:
            # Find matching actual sales
            actual_count = 0
            for db_category, count in actual_sales:
                mapped_category = category_mapping.get(db_category, db_category)
                if mapped_category == category:
                    actual_count = count
                    break
            
            # Find matching target
            target_count = 0
            if target_data:
                if category == 'Classic':
                    target_count = target_data.classic or 0
                elif category == 'Premier Bronze':
                    target_count = target_data.bronze or 0
                elif category == 'Premier Silver':
                    target_count = target_data.silver or 0
                elif category == 'Premier Gold':
                    target_count = target_data.gold or 0
                elif category == 'Premier Platinum':
                    target_count = target_data.platinum or 0
                elif category == 'Safe Card':
                    target_count = target_data.safe_card or 0
                elif category == 'Senior':
                    target_count = target_data.senior or 0
                elif category == 'Senior Plus':
                    target_count = target_data.senior_plus or 0
            
            chart_data.append({
                'category': category,
                'actual': actual_count,
                'target': target_count
            })
        
        return jsonify({
            'success': True,
            'chart_data': chart_data
        })
        
    except Exception as e:
        print(f"Error getting sales performance: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'chart_data': []
        }), 500
    finally:
        db_session.close()

@server.route('/api/dashboard/unit_sold', methods=['GET'])
@login_required
def get_unit_sold():
    """Get unit sold data for the table"""
    try:
        db_session = SessionLocal()
        current_year = datetime.now().year
        
        # Get actual units sold per category
        units_sold = db_session.query(
            models.Entries.maab_category,
            func.count(models.Entries.entry_id).label('units_sold')
        ).filter(
            extract('year', models.Entries.OR_date) == current_year,
            models.Entries.paid == True
        ).group_by(models.Entries.maab_category).all()
        
        # Get targets
        target_data = db_session.query(models.Report_TvA).filter(
            models.Report_TvA.year == current_year
        ).first()
        
        # Price mapping
        price_mapping = {
            'Classic': 60,
            'Bronze': 150,
            'Silver': 300, 
            'Gold': 500,
            'Platinum': 1000,
            'Safe Card': 1200,
            'Senior': 300,
            'Senior+': 350
        }
        
        # Category display mapping
        display_mapping = {
            'Classic': 'Classic',
            'Bronze': 'Premiere Bronze',
            'Silver': 'Premiere Silver',
            'Gold': 'Premiere Gold',
            'Platinum': 'Premiere Platinum',
            'Safe Card': 'Safe Card',
            'Senior': 'Senior',
            'Senior+': 'Senior Plus'
        }
        
        table_data = []
        total_actual = 0
        
        all_categories = ['Classic', 'Bronze', 'Silver', 'Gold', 'Platinum', 'Safe Card', 'Senior', 'Senior+']
        
        for category in all_categories:
            # Find units sold
            units = 0
            for db_category, count in units_sold:
                if db_category == category:
                    units = count
                    break
            
            # Find target
            target = 0
            if target_data:
                if category == 'Classic':
                    target = target_data.classic or 0
                elif category == 'Bronze':
                    target = target_data.bronze or 0
                elif category == 'Silver':
                    target = target_data.silver or 0
                elif category == 'Gold':
                    target = target_data.gold or 0
                elif category == 'Platinum':
                    target = target_data.platinum or 0
                elif category == 'Safe Card':
                    target = target_data.safe_card or 0
                elif category == 'Senior':
                    target = target_data.senior or 0
                elif category == 'Senior+':
                    target = target_data.senior_plus or 0
            
            price = price_mapping.get(category, 0)
            actual_sales = units * price
            total_actual += actual_sales
            
            table_data.append({
                'category': display_mapping.get(category, category),
                'price': price,
                'units_sold': units,
                'target': target,
                'actual_sales': actual_sales
            })
        
        return jsonify({
            'success': True,
            'table_data': table_data,
            'total_actual': total_actual
        })
        
    except Exception as e:
        print(f"Error getting unit sold data: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'table_data': [],
            'total_actual': 0
        }), 500
    finally:
        db_session.close()

@server.route('/api/debug/sales_calculation', methods=['GET'])
@login_required
def debug_sales_calculation():
    """Debug endpoint to see what data we're working with"""
    try:
        db_session = SessionLocal()
        current_year = datetime.now().year
        
        # Get all paid entries for current year
        paid_entries = db_session.query(
            models.Entries.maab_category,
            func.count(models.Entries.entry_id).label('count')
        ).filter(
            extract('year', models.Entries.OR_date) == current_year,
            models.Entries.paid == True
        ).group_by(models.Entries.maab_category).all()
        
        # Get target data
        target_data = db_session.query(models.Report_TvA).filter(
            models.Report_TvA.year == current_year
        ).first()
        
        # Price mapping
        price_map = {
            'Classic': 60,
            'Bronze': 150,
            'Silver': 300,
            'Gold': 500,
            'Platinum': 1000,
            'Safe Card': 1000,
            'Senior': 300,
            'Senior+': 350
        }
        
        # Calculate totals
        total_actual_sales = 0
        category_details = []
        
        for category, count in paid_entries:
            price = price_map.get(category, 0)
            category_total = count * price
            total_actual_sales += category_total
            category_details.append({
                'category': category,
                'count': count,
                'price': price,
                'category_total': category_total
            })
        
        # Calculate target sales
        total_target_sales = 0
        target_details = []
        
        if target_data:
            for category, price in price_map.items():
                target_count = 0
                if category == 'Classic':
                    target_count = target_data.classic or 0
                elif category == 'Bronze':
                    target_count = target_data.bronze or 0
                elif category == 'Silver':
                    target_count = target_data.silver or 0
                elif category == 'Gold':
                    target_count = target_data.gold or 0
                elif category == 'Platinum':
                    target_count = target_data.platinum or 0
                elif category == 'Safe Card':
                    target_count = target_data.safe_card or 0
                elif category == 'Senior':
                    target_count = target_data.senior or 0
                elif category == 'Senior+':
                    target_count = target_data.senior_plus or 0
                
                target_total = target_count * price
                total_target_sales += target_total
                target_details.append({
                    'category': category,
                    'target_count': target_count,
                    'price': price,
                    'target_total': target_total
                })
        
        return jsonify({
            'success': True,
            'debug_info': {
                'current_year': current_year,
                'total_paid_entries': sum([count for _, count in paid_entries]),
                'category_details': category_details,
                'target_details': target_details,
                'total_actual_sales': total_actual_sales,
                'total_target_sales': total_target_sales
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        db_session.close()
        
@server.route('/api/export_record_entries/<int:record_id>')
@login_required
def export_record_entries(record_id):
    try:
        # Fetch record details
        db_session = db_conn.SessionLocal()
        
        record = db_session.query(db_conn.models.Records).filter_by(record_id=record_id).first()
        
        if not record:
            return jsonify({'error': 'Record not found'}), 404
        
        # Fetch entries for this record
        entries = (
            db_session.query(
                db_conn.models.Entries,
                db_conn.models.Members
            )
            .join(db_conn.models.Members, db_conn.models.Entries.member_id == db_conn.models.Members.member_id)
            .filter(db_conn.models.Entries.record_id == record_id)
            .all()
        )
        
        if not entries:
            return jsonify({'error': 'No entries found for this record'}), 404
        
        # Prepare data for Excel with required fields only
        data = []
        for entry, member in entries:
            data.append({
                # Entry details
                'Category': entry.maab_category or '',
                'MAAB No': entry.maab_no or '',
                'First Name': member.first_name or '',
                'Middle Name': member.middle_name or '',
                'Last Name': member.last_name or '',
                'Suffix': member.suffix or 'NA',
                'Birthdate': member.birth_date.strftime('%Y-%m-%d') if member.birth_date else '',
                'Age': member.age or '',
                'Sex': member.sex or '',
                'Contact No': f"+63{member.contact_no}" if member.contact_no else '',
                'Email': member.email or '',
                'OR No': entry.OR_num or '',
                'OR Date': entry.OR_date.strftime('%Y-%m-%d') if entry.OR_date else '',
                'Declared': 'Yes' if entry.declared else 'No',
                'Dispatch Ready': 'Yes' if entry.dispatch_ready else 'No',
                'Dispatch ID': entry.dispatch_id or '',
                'Remarks': entry.remarks or '',
                # Record details (these will be the same for all entries in this record)
                'Declaration Date': record.declaration_date.strftime('%Y-%m-%d') if record.declaration_date else '',
                'Effectivity Date': record.effectivity_date.strftime('%Y-%m-%d') if record.effectivity_date else '',
                'Location Category': record.location_category or '',
                'Municipality': record.municipality or '',
                'District': record.district or '',
                'Origin': record.origin or ''
            })
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create the record_exports folder if it doesn't exist
        export_folder = os.path.join(os.path.dirname(__file__), 'record_exports')
        os.makedirs(export_folder, exist_ok=True)
        
        # Generate filename
        loc_particular = record.location_particular or 'Unknown_Location'
        dec_date = record.declaration_date.strftime('%Y-%m-%d') if record.declaration_date else 'no_date'
        
        # Clean filename (remove special characters)
        clean_location = "".join(c for c in loc_particular if c.isalnum() or c in (' ', '-', '_')).rstrip()
        clean_location = clean_location.replace(' ', '_')
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{clean_location}_{dec_date}_{timestamp}.xlsx"
        file_path = os.path.join(export_folder, filename)
        
        # Create Excel file and save to server folder
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write the main title first
            workbook = writer.book
            worksheet = workbook.create_sheet('Entries')
            
            # Add main title
            worksheet.merge_cells('A1:W1')
            worksheet['A1'] = f"{loc_particular} - {dec_date}"
            worksheet['A1'].font = openpyxl.styles.Font(size=16, bold=True)
            worksheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Add record details as separate rows
            record_details = [
                f"Declaration Date: {record.declaration_date.strftime('%Y-%m-%d') if record.declaration_date else 'N/A'}",
                f"Effectivity Date: {record.effectivity_date.strftime('%Y-%m-%d') if record.effectivity_date else 'N/A'}",
                f"Location Category: {record.location_category or 'N/A'}",
                f"Municipality: {record.municipality or 'N/A'}",
                f"District: {record.district or 'N/A'}",
                f"Origin: {record.origin or 'N/A'}"
            ]
            
            # Write record details starting from row 2
            for i, detail in enumerate(record_details, start=2):
                worksheet[f'A{i}'] = detail
            
            # Write the DataFrame starting from row 8 (after title and record details)
            start_row = 8
            # Write headers
            for col_num, column_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=start_row, column=col_num)
                cell.value = column_name
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Write data rows
            for row_num, row_data in enumerate(df.values, start_row + 1):
                for col_num, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col_num, value=cell_value)
            
            # Adjust column widths for all required columns
            column_widths = {
                'A': 15,  # Category
                'B': 15,  # MAAB No
                'C': 15,  # First Name
                'D': 15,  # Middle Name
                'E': 15,  # Last Name
                'F': 8,   # Suffix
                'G': 12,  # Birthdate
                'H': 6,   # Age
                'I': 8,   # Sex
                'J': 15,  # Contact No
                'K': 20,  # Email
                'L': 10,  # OR No
                'M': 10,  # OR Date
                'N': 10,  # Declared
                'O': 15,  # Dispatch Ready
                'P': 12,  # Dispatch ID
                'Q': 20,  # Remarks
                'R': 12,  # Declaration Date (record)
                'S': 12,  # Effectivity Date (record)
                'T': 20,  # Location Category (record)
                'U': 15,  # Municipality (record)
                'V': 10,  # District (record)
                'W': 12   # Origin (record)
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
        
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Export Record', f'Exported record {record_id} with {len(entries)} entries', current_user.account_id)
        return jsonify({
            'success': True,
            'message': f'File exported successfully to record_exports folder',
            'filename': filename,
            'file_path': file_path,
            'export_count': len(entries)
        })
        
    except Exception as e:
        print(f"Export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Internal server error'}), 500
    finally:
        db_session.close()

# ========================== FORGOT PASSWORD ==========================
@server.route("/forgot_password_otp", methods=["GET", "POST"])
def forgot_password_otp():
    if request.method == "GET":
        return render_template("forgot_password.html")
    
    if request.method == "POST":
        try:
            # Get email from form
            email = request.form.get("email", "").strip()
            print(f"🔍 DEBUG: Starting forgot password process for: '{email}'")
            
            # Validate email format
            if not email:
                flash({
                    "title": "Email Required",
                    "text": "Please enter your email address.",
                    "redirect_url": url_for('forgot_password_otp')
                }, "error")
                return render_template("forgot_password.html")
            
            if not tools.validate_email_format(email):
                flash({
                    "title": "Invalid Email",
                    "text": "Please enter a valid email address.",
                    "redirect_url": url_for('forgot_password_otp')
                }, "error")
                return render_template("forgot_password.html")

            # Check if email exists in accounts table and user is authorized
            user_exists = db_conn.check_user_exists(email)
            print(f"🔍 DEBUG: User exists: {user_exists}")
            
            if not user_exists:
                flash({
                    "title": "Email Not Found",
                    "text": "This email is not registered in our system.",
                    "redirect_url": url_for('forgot_password_otp')
                }, "error")
                return render_template("forgot_password.html")
            
            # Check if user is authorized (staff or allowed for OTP)
            user_authorized = db_conn.check_user_authorized(email)
            print(f"🔍 DEBUG: User authorized: {user_authorized}")
            
            if not user_authorized:
                flash({
                    "title": "Not Authorized",
                    "text": "Your account is not authorized for password reset via OTP. Please contact support.",
                    "redirect_url": url_for('forgot_password_otp')
                }, "error")
                return render_template("forgot_password.html")

            # Generate OTP
            otp = tools.generate_otp()
            print(f"🔍 DEBUG: Generated OTP: {otp}")
            
            # Save OTP in database
            save_otp = db_conn.save_otp(email, otp)
            print(f"🔍 DEBUG: Save OTP result: {save_otp}")
            
            if not save_otp:
                flash({
                    "title": "OTP Error",
                    "text": "Failed to generate OTP. Please try again.",
                    "redirect_url": url_for('forgot_password_otp')
                }, "error")
                return render_template("forgot_password.html")

            # Send OTP email
            print(f"🔍 DEBUG: Attempting to send OTP email to: {email}")
            email_sent = db_conn.send_otp_email(email, otp)
            print(f"🔍 DEBUG: Email sent result: {email_sent}")
            
            if not email_sent:
                print(f"❌ DEBUG: Email sending failed for: {email}")
                flash({
                    "title": "Email Error",
                    "text": "Failed to send OTP email. Please try again later.",
                    "redirect_url": url_for('forgot_password_otp')
                }, "error")
                return render_template("forgot_password.html")

            # Store email in session for verification
            session["email"] = email
            session["otp_verified"] = False

            print(f"✅ DEBUG: SUCCESS! OTP process completed for: {email}")
            print(f"✅ DEBUG: Redirecting to verify_otp page")
            
            session.pop('_flashes', None)
            # Success
            flash({
                "title": "OTP Sent!",
                "text": "OTP has been sent to your email.",
                "redirect_url": url_for('verify_otp')
            }, "info")
            return redirect(url_for('verify_otp'))

        except Exception as e:
            print(f"❌ ERROR in forgot_password_otp: {e}")
            import traceback
            traceback.print_exc()
            flash({
                "title": "System Error",
                "text": "An error occurred. Please try again.",
                "redirect_url": url_for('forgot_password_otp')
            }, "error")
            return render_template("forgot_password.html")


# ========================== VERIFY OTP ==========================
@server.route('/verify_otp', methods=['GET', 'POST'])
def verify_otp():
    # Check if user came from forgot password flow
    email = session.get("email")
    if not email:
        flash({
            "title": "Session Expired",
            "text": "Please start the password reset process again.",
            "redirect_url": url_for('forgot_password_otp')
        }, "error")
        return redirect(url_for('forgot_password_otp'))

    if request.method == 'GET':
        return render_template('verify_otp.html', email=email)

    if request.method == 'POST':
        try:
            otp_input = request.form.get("otp", "").strip()

            if not otp_input:
                flash({
                    "title": "OTP Required",
                    "text": "Please enter the OTP sent to your email.",
                    "redirect_url": url_for('verify_otp')
                }, "error")
                return render_template('verify_otp.html', email=email)

            # Verify OTP
            result = db_conn.verifying_otp(email, otp_input)

            if result == "success":
                session["otp_verified"] = True
                flash({
                    "title": "OTP Verified Successfully!",
                    "text": "You can now proceed to reset your password.",
                    "redirect_url": url_for('reset_password')
                }, "success")
                return redirect(url_for('reset_password'))

            elif result == "expired":
                flash({
                    "title": "OTP Expired",
                    "text": "OTP has expired. Please request a new one.",
                    "redirect_url": url_for('forgot_password_otp')
                }, "error")
            elif result == "already_used":
                flash({
                    "title": "OTP Used",
                    "text": "OTP has already been used. Please request a new one.",
                    "redirect_url": url_for('forgot_password_otp')
                }, "error")
            elif result == "email_not_found":
                flash({
                    "title": "Email Error",
                    "text": "Email verification failed. Please start over.",
                    "redirect_url": url_for('forgot_password_otp')
                }, "error")
            else:
                flash({
                    "title": "Invalid OTP",
                    "text": "The OTP you entered is incorrect. Please try again.",
                    "redirect_url": url_for('verify_otp')
                }, "error")

            return render_template('verify_otp.html', email=email)

        except Exception as e:
            print(f"Error in verify_otp: {e}")
            flash({
                "title": "System Error",
                "text": "An error occurred. Please try again.",
                "redirect_url": url_for('verify_otp')
            }, "error")
            return render_template('verify_otp.html', email=email)


# ========================== RESET PASSWORD ==========================
@server.route("/reset_password", methods=["GET", "POST"])
def reset_password():
    # Check if OTP was verified
    email = session.get("email")
    otp_verified = session.get("otp_verified")
    
    if not email or not otp_verified:
        flash({
            "title": "Session Expired",
            "text": "Please complete OTP verification first.",
            "redirect_url": url_for('forgot_password_otp')
        }, "error")
        print("🔍 DEBUG: Flashed session expired message")
        return redirect(url_for('forgot_password_otp'))

    if request.method == "GET":
        return render_template("reset_password.html")

    if request.method == "POST":
        try:
            new_password = request.form.get("new_password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()

            # Validate passwords
            if not new_password or not confirm_password:
                flash("Please fill in all password fields.", "error")
                print("🔍 DEBUG: Flashed empty fields error")
                return render_template("reset_password.html")

            if new_password != confirm_password:
                flash("Passwords do not match. Please try again.", "error")
                print("🔍 DEBUG: Flashed password mismatch error")
                return render_template("reset_password.html")

            # Update password
            success = db_conn.update_password(email, new_password)
            
            if success:
                # Clear session
                session.pop("email", None)
                session.pop("otp_verified", None)
                
                flash({
                    "title": "Password Reset Successfully!",
                    "text": "You can now log in with your new password.",
                    "redirect_url": url_for('landing_page')
                }, "success")
                print("🔍 DEBUG: Flashed success message")
                return redirect(url_for('landing_page'))
            else:
                flash("Failed to reset password. Please try again.", "error")
                print("🔍 DEBUG: Flashed reset failed error")
                return render_template("reset_password.html")
                
        except Exception as e:
            print(f"Error in reset_password: {e}")
            flash("An error occurred while resetting password. Please try again.", "error")
            print("🔍 DEBUG: Flashed exception error")
            return render_template("reset_password.html")

# KEEP ONLY THIS VERSION OF THE ROUTE - REMOVE THE DUPLICATE LATER IN THE FILE
@server.route('/api/declaration', methods=['POST'])
def declaration_api():
    try:
        data = request.get_json()
        print(f"📦 DECLARATION API CALLED - Data: {data}")  # Debug print
        
        if not data:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Dispatch Failed', 'No data provided', current_user.account_id)
            return jsonify({"success": False, "error": "No data provided"}), 400

        dispatch_type = 'transmission' if current_user.office_location != 'Chapter' else 'declaration' 
        dispatch_origin = current_user.office_location
        dispatch_year = datetime.now().year
        dispatch_cutoff = data.get('dispatch_cutoff')
        late_declare = data.get('late_declare')
        dispatch_remarks = data.get('dispatch_remarks')
        
        print(f"📦 Creating dispatch with: type={dispatch_type}, origin={dispatch_origin}, "
              f"year={dispatch_year}, cutoff={dispatch_cutoff}, late_declare={late_declare}")
        
        result = db_conn.create_dispatch(dispatch_type, dispatch_origin, dispatch_year, dispatch_cutoff, late_declare, dispatch_remarks)
        
        print(f"📦 DB Result: {result}")  # Debug print
        
        if result == True:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Dispatch', f'Created {dispatch_type} dispatch from {dispatch_origin}', current_user.account_id)
            return jsonify({"success": True, "message": "Dispatch created successfully"})
            
        else:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Dispatch Failed', f'Failed: {result}', current_user.account_id)
            return jsonify({"success": False, "error": result}), 500
            
    except Exception as e:
        print(f"❌ Declaration API error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": "Internal server error"}), 500
    
@server.route('/settings_save_changes', methods=['POST'])
@login_required
def settings_save_changes():
    if request.method == 'POST':
        print("=== FORM DATA RECEIVED ===")
        print(request.form)
        print("===========================")

        first_name = request.form.get('first_name', '').upper()
        middle_name = request.form.get('middle_name', '').upper()
        last_name = request.form.get('last_name', '').upper()
        birthdate = request.form.get('birthdate')
        email = request.form.get('email')
        phone = request.form.get('phone')

        if all([first_name, middle_name, last_name, birthdate, email, phone]):
            update_success = db_conn.update_user_details(
            current_user.account_id,  # ✅ actual user ID
            first_name,
            middle_name,
            last_name,
            birthdate,
            phone,
            email
            )

            if update_success:
                db_conn.POST_action_log(current_user.username, current_user.user_level, 'Update Profile', 'Updated personal profile information', current_user.account_id)
                flash("Details updated successfully!", "success")
            else:
                flash("Update failed!", "error")

        else:
            flash("Please fill out all fields.", "error")

        return redirect(url_for('settings'))

    return redirect(url_for('settings'))

# === Upload profile picture ===
@server.route('/api/upload_profile_pic', methods=['POST'])
@login_required
def upload_profile_pic():
    try:
        # Debug print
        print("=== /api/upload_profile_pic called ===")
        if 'profile_pic' not in request.files:
            print("No 'profile_pic' in request.files. Keys:", request.files.keys())
            return jsonify({"success": False, "error": "No file part 'profile_pic'"}), 400

        file = request.files['profile_pic']
        if file.filename == '':
            print("Empty filename")
            return jsonify({"success": False, "error": "Empty filename"}), 400

        file_data = file.read()
        print(f"Received file: name={file.filename}, size={len(file_data)} bytes, content_type={file.content_type}")

        # Call DB helper to save the profile pic
        saved = db_conn.save_profile_pic(current_user.account_id, file_data)
        if saved:
            print("Saved profile pic to DB for user", current_user.account_id)
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Upload Profile Picture', 'Uploaded new profile picture', current_user.account_id)
            return jsonify({"success": True, "has_profile_pic": True})
        else:
            print("db_conn.save_profile_pic returned False")
            return jsonify({"success": False, "error": "DB save failed"}), 500

    except Exception as e:
        print("Exception in upload_profile_pic:", e)
        return jsonify({"success": False, "error": str(e)}), 500

@server.route('/api/members/filter_options', methods=['GET'])
@login_required
def get_member_filter_options():
    db_session = SessionLocal()
    try:
        # Get distinct years from records
        years = db_session.query(models.Records.year).distinct().filter(models.Records.year.isnot(None)).order_by(models.Records.year.desc()).all()
        years_list = [year[0] for year in years]
        
        # Get distinct origins from records
        origins = db_session.query(models.Records.origin).distinct().filter(models.Records.origin.isnot(None)).order_by(models.Records.origin).all()
        origins_list = [origin[0] for origin in origins]
        
        # Ensure current user's office location is in the origins list
        user_location = current_user.office_location if current_user else None
        if user_location and user_location not in origins_list:
            origins_list.append(user_location)
            origins_list.sort()
        
        return jsonify({
            'years': years_list,
            'origins': origins_list
        })
        
    except Exception as e:
        print(f"Error fetching filter options: {e}")
        return jsonify({'years': [], 'origins': []}), 500
    finally:
        db_session.close()

# === Fetch profile picture ===
@server.route('/api/get_profile_pic/<int:user_id>')
@login_required
def get_profile_pic(user_id):
    image_data = db_conn.get_profile_pic(user_id)
    if image_data:
        return send_file(BytesIO(image_data), mimetype='image/jpeg')
    else:
        return send_file('static/assets/pfp.jpg', mimetype='image/jpeg')

# === Inject profile picture state into templates ===
@server.context_processor
def inject_profile_pic_state():
    if current_user.is_authenticated:
        has_pic = db_conn.get_profile_pic(current_user.account_id) is not None
        return {"has_profile_pic": has_pic}
    return {"has_profile_pic": False}

# === Delete profile picture ===
@server.route('/api/delete_profile_pic', methods=['DELETE'])
@login_required
def delete_profile_pic():
    success = db_conn.save_profile_pic(current_user.account_id, None)
    db_conn.POST_action_log(current_user.username, current_user.user_level, 'Delete Profile Picture', 'Deleted profile picture', current_user.account_id)
    return jsonify({"success": success})

# API FOR DASHBOARD

# ====== API FOR DASHBOARD STATS ======

@server.route('/api/members/count', methods=['GET'])
def get_members_count():
    try:
        # Count all entries across all records
        total_count = 0
        all_records = db_conn.get_member_records()
        
        for record in all_records:
            entries = db_conn.get_entries(record.record_id)  # or record.id
            total_count += len(entries)
            
        return jsonify({
            'success': True,
            'total_members': total_count
        })
    except Exception as e:
        print(f"Error in get_members_count: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@server.route('/api/members/expiring_soon', methods=['GET'])
def get_expiring_soon_count():
    try:
        from datetime import datetime, timedelta
        today = datetime.now().date()
        thirty_days_from_now = today + timedelta(days=30)
        last_month = today - timedelta(days=30)
        
        # Get all records and their entries
        all_records = db_conn.get_member_records()
        
        current_expiring = 0
        previous_expiring = 0
        
        for record in all_records:
            # Get entries for this record
            entries = db_conn.get_entries(record.record_id)
            
            for entry in entries:
                # Check OR_date for expiration
                if hasattr(entry, 'OR_date') and entry.OR_date:
                    # Convert to date object if it's string
                    if isinstance(entry.OR_date, str):
                        or_date = datetime.strptime(entry.OR_date, '%Y-%m-%d').date()
                    else:
                        or_date = entry.OR_date
                    
                    # Calculate expiration date (OR_date + 1 year)
                    expiration_date = or_date + timedelta(days=365)
                    
                    # Current period: expiring in next 30 days
                    if today <= expiration_date <= thirty_days_from_now:
                        current_expiring += 1
                    
                    # Previous period: expired in last 30 days
                    if last_month <= expiration_date <= today:
                        previous_expiring += 1
        
        # Calculate percentage change
        if previous_expiring > 0:
            percentage_change = ((current_expiring - previous_expiring) / previous_expiring) * 100
        else:
            percentage_change = 0
        
        return jsonify({
            'success': True,
            'expiring_soon_count': current_expiring,
            'previous_period_count': previous_expiring,
            'percentage_change': round(percentage_change, 2),
            'timeframe_days': 30
        })
    except Exception as e:
        print(f"Error in get_expiring_soon_count: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    
@server.route('/api/get_pending_claims_count')
def get_pending_claims_count():
    count = db_conn.get_pending_claims_count()
    return jsonify({"pending_claims_count": count})


@server.route('/api/get_all_dispatch')
def get_dispatch_records():
    dispatch_records = db_conn.get_all_dispatch_records()
    if dispatch_records is None:
        return jsonify({"success": False, "error": "Failed to fetch dispatch records"}), 500
    print('flask_server: dispatch_records', [record.to_dict() for record in dispatch_records])
    return jsonify([record.to_dict() for record in dispatch_records])


# API FOR MEMBERS PAGE

# TODO fix api route names ex. /api/members/get_records

@server.route('/api/members/records', methods=['GET'])
def get_members():
    try:
        status = request.args.get('status', 'active')
        office_loc = request.args.get('office_loc', None)
        
        print(f"DEBUG: Fetching member records with status: {status}, office_loc: {office_loc}")
        
        member_records = db_conn.get_member_records(status=status, office_loc=office_loc)
        print(f"DEBUG: Records fetched: {len(member_records) if member_records else 'None'}")
        
        if member_records is None:
            print("DEBUG: No records returned from database")
            return jsonify({"error": "Failed to fetch member records"}), 500
            
        records_list = []
        for record in member_records:
            try:
                record_dict = record.to_dict()
                
                # ADD ENTRY COUNT TO EACH RECORD
                record_id = record_dict.get('record_id')
                if record_id:
                    # Get entry count for this record
                    entries = db_conn.get_entries(record_id)
                    record_dict['entries_count'] = len(entries) if entries else 0
                else:
                    record_dict['entries_count'] = 0
                    
                records_list.append(record_dict)
            except Exception as e:
                print(f"DEBUG: Error converting record to dict: {e}")
                continue
                
        print(f"DEBUG: Returning {len(records_list)} records")
        return jsonify(records_list)
        
    except Exception as e:
        print(f"DEBUG: Error in get_members: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Internal server error"}), 500
    '''
    records = db_conn.get_member_records()
    members_list = []

    for row in records:
        record = dict(row._mapping)

        for key in ['id_received', 'declared', 'paid']:
            if key in record:
                record[key] = bool(record[key])

        for key in ['declaration_date', 'effectivity_date']:
            if key in record and record[key]:
                record[key] = record[key].strftime('%Y-%m-%d')

        members_list.append(record)

    return jsonify(members_list)
    '''
# ============================================================
# 🔹 DEBUG ROUTES FOR MAAB NUMBER ISSUES
# ============================================================

@server.route('/api/debug/maab_formats', methods=['POST'])
@login_required
def debug_maab_formats():
    """Debug route to check MAAB number formats"""
    try:
        data = request.get_json()
        maab_numbers = data.get('maab_numbers', [])
        
        import re
        results = {}
        
        for maab_no in maab_numbers:
            # Current strict validation (7 digits)
            strict_match = re.match(r'^(PC|PB|PS|PG|PP|PEP|S|SP)\d{7}$', maab_no)
            
            # Flexible validation (any digits)
            flexible_match = re.match(r'^(PC|PB|PS|PG|PP|PEP|S|SP)\d+$', maab_no)
            
            results[maab_no] = {
                'strict_validation': bool(strict_match),
                'flexible_validation': bool(flexible_match),
                'length': len(maab_no),
                'prefix': maab_no[:2] if len(maab_no) >= 2 else 'N/A',
                'numbers': maab_no[2:] if len(maab_no) > 2 else 'N/A'
            }
        
        return jsonify({
            'success': True,
            'validation_results': results,
            'note': 'Your MAAB numbers should pass flexible_validation to work with the fixed route'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# ============================================================
# 🔹 GET USED MAAB NUMBERS FROM ENTRIES
# ============================================================

@server.route('/api/members/used_maab_numbers', methods=['GET'])
@login_required
def get_used_maab_numbers():
    """Get all MAAB numbers that are already used in entries"""
    try:
        db_session = SessionLocal()
        
        try:
            # Query all used MAAB numbers from entries table
            cursor = db_session.execute(text("""
                SELECT DISTINCT maab_no 
                FROM entry_contents 
                WHERE maab_no IS NOT NULL AND maab_no != ''
            """))
            used_numbers = [row[0] for row in cursor.fetchall()]
            
            print(f"🔍 Found {len(used_numbers)} used MAAB numbers in entries")
            
            return jsonify({
                'success': True,
                'used_numbers': used_numbers,
                'count': len(used_numbers)
            })
            
        except Exception as e:
            print(f"❌ Database error in get_used_maab_numbers: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': 'Database error occurred'
            }), 500
            
        finally:
            db_session.close()
            
    except Exception as e:
        print(f"❌ Error in get_used_maab_numbers route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500

@server.route('/api/archive_record', methods=['PATCH'])
@login_required
def archive_record():
    try:
        data = request.get_json()
        print(f"🔍 ARCHIVE RECORD API CALLED - Data: {data}")
        
        if not data:
            return jsonify({"success": False, "error": "No data provided"}), 400
        
        record_id = data.get('record_id')
        
        if not record_id:
            return jsonify({"success": False, "error": "No record ID provided"}), 400
        
        print(f"🔄 Archiving record: {record_id}")

        # Use the archive function
        success = db_conn.archive_member_record_with_log(record_id, current_user.account_id)
        
        if success:
            print(f"✅ Record {record_id} archived successfully")
            return jsonify({
                "success": True, 
                "message": "Record archived successfully"
            })
        else:
            print(f"❌ Failed to archive record {record_id}")
            return jsonify({
                "success": False, 
                "error": "Failed to archive record - record not found or already archived"
            }), 500
            
    except Exception as e:
        print(f"❌ Error archiving record: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False, 
            "error": f"Internal server error: {str(e)}"
        }), 500
    
@server.route('/api/debug/archive_check/<int:record_id>', methods=['GET'])
@login_required
def debug_archive_check(record_id):
    """Debug route to check if a record can be archived"""
    try:
        db_session = SessionLocal()
        
        # Check if record exists
        record = db_session.query(models.Records).filter_by(record_id=record_id).first()
        
        if not record:
            return jsonify({
                "success": False,
                "error": f"Record {record_id} not found",
                "exists": False
            })
        
        return jsonify({
            "success": True,
            "record_exists": True,
            "record_id": record.record_id,
            "current_status": record.status,
            "can_be_archived": record.status != 'archived'
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500
    finally:
        db_session.close()
    
@server.route('/api/add_record', methods=['POST'])
def add_new_record():
    new_record_id = db_conn.add_new_record()
    return jsonify({"success": True, "record_id": new_record_id})

# Verify your existing save_record_details route has this structure:
@server.route('/api/save_record_details', methods=['POST', 'PATCH'])
def save_record_details():
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "error": "No data provided"}), 400
    
    db_session = SessionLocal()
    try:
        if request.method == 'POST':
            # Create new record - record_id should be None/Null for auto-increment
            new_record = models.Records(
                year=data.get('year', datetime.now().year),
                id_received=data.get('id_received', 0),
                declared=data.get('declared', 0),
                declaration_date=data.get('declaration_date'),
                effectivity_date=data.get('effectivity_date'),
                location_particular=data.get('location_particular'),
                location_category=data.get('location_category'),
                municipality=data.get('municipality'),
                district=data.get('district'),
                paid=data.get('paid', 0),
                origin=data.get('origin'),
                remarks=data.get('remarks'),
                tags=data.get('tags'),
                dispatch_ready=data.get('dispatch_ready', 0),
                status='active'
            )
            db_session.add(new_record)
            db_session.commit()
            
            return jsonify({
                "success": True, 
                "record_id": new_record.record_id,  # This returns the actual auto-generated ID
                "message": "Record created successfully"
            })
            
        else:  # PATCH method for updates
            record_id = data.get('record_id')
            if not record_id:
                return jsonify({"success": False, "error": "No record ID provided"}), 400
                
            record = db_session.query(models.Records).filter_by(record_id=record_id).first()
            if not record:
                return jsonify({"success": False, "error": "Record not found"}), 404

            # Update fields if present in data
            update_fields = [
                'year', 'id_received', 'declared', 'declaration_date', 'effectivity_date',
                'location_particular', 'location_category', 'municipality', 'district',
                'paid', 'origin', 'remarks', 'tags', 'dispatch_ready'
            ]
            
            for field in update_fields:
                if field in data:
                    value = data[field]
                    # Convert empty string to None for date/string fields
                    if value == '':
                        value = None
                    setattr(record, field, value)
                    
            db_session.commit()
            return jsonify({"success": True, "message": "Record updated successfully"})
            
    except Exception as e:
        db_session.rollback()
        print(f"Error saving record details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        db_session.close()


@server.route('/api/members/<int:record_id>/entries', methods=['GET'])
def get_entries(record_id):
    entries = db_conn.get_entries(record_id)
    return jsonify(entries)

@server.route('/api/save_entry_details', methods=['POST'])
def save_entry_details():
    data = request.get_json()
    print("=== ENTRY SAVE REQUEST ===")
    print("Received data:", data)
    
    if not data:
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Add Entry Failed', 'No data provided', current_user.account_id)
        return jsonify({"success": False, "error": "No data provided"}), 400
    
    db_session = SessionLocal()
    try:
        # Extract and validate required fields
        record_id = data.get('record_id')
        if not record_id:
            return jsonify({"success": False, "error": "Record ID is required"}), 400

        # Extract member data with proper None handling
        first_name = data.get('first_name')
        middle_name = data.get('middle_name')
        last_name = data.get('last_name')
        suffix = data.get('suffix', 'NA')
        
        # Validate required fields
        if not first_name or not last_name:
            return jsonify({"success": False, "error": "First name and last name are required"}), 400

        # Handle string fields - convert None to empty string, then strip
        first_name = (first_name or '').strip().upper()
        middle_name = (middle_name or '').strip().upper()
        last_name = (last_name or '').strip().upper()
        
        # Parse birthdate
        birthdate = None
        birthdate_string = data.get('birth_date')
        if birthdate_string:
            try:
                birthdate = datetime.strptime(birthdate_string, "%Y-%m-%d").date()
            except ValueError as e:
                print(f"Birthdate parsing error: {e}")
                # Continue without birthdate rather than failing

        # Parse other dates
        declaration_date = None
        declaration_date_string = data.get('declaration_date')
        if declaration_date_string:
            try:
                declaration_date = datetime.strptime(declaration_date_string, "%Y-%m-%d").date()
            except ValueError:
                pass

        OR_date = None
        OR_date_string = data.get('OR_date')
        if OR_date_string:
            try:
                OR_date = datetime.strptime(OR_date_string, "%Y-%m-%d").date()
            except ValueError:
                pass

        # Create new member with proper None handling for all fields
        new_member = models.Members(
            first_name=first_name,
            middle_name=middle_name or None,  # Convert empty string back to None
            last_name=last_name,
            suffix=suffix,
            birth_date=birthdate,
            age=data.get('age'),
            sex=data.get('sex'),
            contact_no=data.get('contact_no'),
            email=data.get('email'),
            address=data.get('address'),
            blood_type=data.get('blood_type')
        )
        db_session.add(new_member)
        db_session.flush()  # Get member_id without committing
        
        member_id = new_member.member_id
        print(f"Created new member with ID: {member_id}")

        # Create new entry
        new_entry = models.Entries(
            record_id=record_id,
            maab_category=data.get('maab_category', 'Classic'),
            maab_no=data.get('maab_no'),
            member_id=member_id,
            id_received=bool(data.get('id_received', False)),
            declared=bool(data.get('declared', False)),
            declaration_date=declaration_date,
            paid=bool(data.get('paid', False)),
            OR_num=data.get('OR_num'),
            OR_date=OR_date,
            remarks=data.get('remarks'),
            tags=data.get('tags'),
            dispatch_ready=bool(data.get('dispatch_ready', False))
        )
        db_session.add(new_entry)
        db_session.flush()
        
        entry_id = new_entry.entry_id
        print(f"Created new entry with ID: {entry_id}")

        # Commit both member and entry
        db_session.commit()
        
        print("=== ENTRY SAVE SUCCESS ===")
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Add Entry', f'Added entry for {first_name} {last_name}', current_user.account_id)
        return jsonify({
            "success": True, 
            "message": "Entry added successfully",
            "entry_id": entry_id,
            "member_id": member_id
        })
        
    except Exception as e:
        db_session.rollback()
        print(f"=== ENTRY SAVE ERROR ===")
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Add Entry Failed', f'Failed to add entry for {first_name} {last_name}', current_user.account_id)
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500
    finally:
        db_session.close()


@server.route('/api/save_entry_update', methods=['POST'])
@login_required
def save_entry_update():
    try:
        data = request.get_json()
        print("=== ENTRY UPDATE API CALL ===")
        print(f"Request data: {data}")
        print(f"Data types: { {k: type(v) for k, v in data.items()} }")
        
        if not data:
            print("❌ No data provided")
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Update Entry Failed', 'No data provided', current_user.account_id)
            return jsonify({"success": False, "error": "No data provided"}), 400
            
        entry_id = data.get('entry_id')
        print(f'Processing update for entry_id: {entry_id}')
        
        if not entry_id:
            print("❌ No entry_id provided")
            return jsonify({"success": False, "error": "No entry ID provided"}), 400
        
        
        # Test database connection first
        print("🔍 Testing database connection...")
        try:
            db_session = SessionLocal()
            test_entry = db_session.query(models.Entries).filter_by(entry_id=entry_id).first()
            if not test_entry:
                print(f"❌ Entry {entry_id} not found in database")
                return jsonify({"success": False, "error": f"Entry {entry_id} not found"}), 404
            print(f"✅ Database connection test passed - Entry found")
            db_session.close()
        except Exception as db_error:
            print(f"❌ Database connection test failed: {db_error}")
            return jsonify({"success": False, "error": f"Database connection failed: {str(db_error)}"}), 500
        
        print("🔍 Calling db_conn.save_entry_updates...")
        result = db_conn.save_entry_update(data)
        print(f"🔍 db_conn.save_entry_updates returned: {result} (type: {type(result)})")
        
        # FIXED: Check for explicit True/False instead of truthy/falsy
        if result is True:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Update Entry', f'Updated entry ID: {entry_id}', current_user.account_id)

            print("✅ Entry update successful")
            return jsonify({"success": True})
        else:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Update Entry Failed', f'Failed to update entry ID: {entry_id}', current_user.account_id)
            print("❌ Entry update failed in db_conn")
            return jsonify({"success": False, "error": "Database update failed"}), 500
            
    except Exception as e:
        print(f"💥 Exception in save_entry_update route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": f"Server error: {str(e)}"}), 500


@server.route('/api/get_report/target_vs_actual/<int:year>', methods=['GET'])
def target_vs_actual(year):
    if not year:
        return jsonify({"success": False, "error": "Year parameter is required"}), 400

    report_data = db_conn.get_report_target_vs_actual(year)
    if report_data is None:
        return jsonify({"success": False, "error": "Failed to fetch report data"}), 500

    return jsonify(report_data)

# =============================================
# REPORTS API ROUTES
# =============================================

@server.route('/api/get_report/target_vs_actual/<year>')
@login_required
@roles_required('admin', 'superadmin')
def get_target_vs_actual(year):
    """Get target vs actual data for a specific year"""
    try:
        db_session = SessionLocal()
        
        # Query the database for targets
        target_data = db_session.query(models.Report_TvA).filter(
            models.Report_TvA.year == year
        ).first()
        
        if target_data:
            # Return actual data from database
            return jsonify({
                "Classic": {0: target_data.classic or 0},
                "Bronze": {0: target_data.bronze or 0},
                "Silver": {0: target_data.silver or 0},
                "Gold": {0: target_data.gold or 0},
                "Platinum": {0: target_data.platinum or 0},
                "Safe Card": {0: target_data.safe_card or 0},
                "Senior": {0: target_data.senior or 0},
                "Senior+": {0: target_data.senior_plus or 0}
            })
        else:
            # Return zeros if no targets set
            return jsonify({
                "Classic": {0: 0},
                "Bronze": {0: 0},
                "Silver": {0: 0},
                "Gold": {0: 0},
                "Platinum": {0: 0},
                "Safe Card": {0: 0},
                "Senior": {0: 0},
                "Senior+": {0: 0}
            })
            
    except Exception as e:
        print(f"Error getting target vs actual data: {e}")
        # Fallback to sample data if error
        sample_data = {
            "Classic": {0: 0}, "Bronze": {0: 0}, "Silver": {0: 0}, "Gold": {0: 0},
            "Platinum": {0: 0}, "Safe Card": {0: 0}, "Senior": {0: 0}, "Senior+": {0: 0}
        }
        return jsonify(sample_data)
    finally:
        db_session.close()

@server.route('/api/get_target_years')
@login_required
@roles_required('admin', 'superadmin')
def get_target_years():
    """Get all years that have target data"""
    try:
        db_session = SessionLocal()
        
        # Get distinct years from target_per_year table
        years = db_session.query(models.Report_TvA.year).distinct().order_by(models.Report_TvA.year.desc()).all()
        
        years_list = [year[0] for year in years]
        
        print(f"📅 Years with target data: {years_list}")
        return jsonify(years_list)
        
    except Exception as e:
        print(f"Error getting target years: {e}")
        return jsonify([])
    finally:
        db_session.close()        

@server.route('/api/get_report/budget_expenses/<year>')
@login_required
@roles_required('admin', 'superadmin')
def get_budget_expenses(year):
    """Get budget vs expenses data for a specific year"""
    try:
        # Sample budget data - replace with actual database queries
        sample_data = [
            {
                "id": 1,
                "account_code": "5020401",
                "account_name": "Gasoline & Oil",
                "budget_2025": 100800,
                "jan": 1400, "feb": 4100, "mar": 3400, "apr": 3700, "may": 2000, "jun": 2100,
                "jul": 1400, "aug": 0, "sep": 0, "oct": 0, "nov": 0, "dec": 0,
                "total_expense": 18100,
                "balance": 82700,
            },
            {
                "id": 2,
                "account_code": "5020511",
                "account_name": "Travel Local",
                "budget_2025": 20000,
                "jan": 0, "feb": 0, "mar": 0, "apr": 0, "may": 60, "jun": 20,
                "jul": 0, "aug": 0, "sep": 0, "oct": 0, "nov": 0, "dec": 0,
                "total_expense": 80,
                "balance": 19920,
            }
        ]
        return jsonify(sample_data)
    except Exception as e:
        print(f"Error getting budget expenses data: {e}")
        return jsonify({"error": str(e)}), 500

@server.route('/api/get_report/per_district/<year>')
@login_required
@roles_required('admin', 'superadmin')
def get_per_district(year):
    """Get per district data for a specific year"""
    try:
        # Return empty array for now - frontend will use sample data
        # You can implement actual database queries here later
        return jsonify([])
    except Exception as e:
        print(f"Error getting per district data: {e}")
        return jsonify({"error": str(e)}), 500

@server.route('/api/save_report/target_vs_actual', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def save_target_vs_actual():
    """Save target vs actual data - Auto-creates year if needed"""
    try:
        data = request.get_json()
        category = data.get('category')
        target_count = data.get('targetCount')
        year = data.get('year')
        
        # Validate required fields
        if not all([category, target_count, year]):
            return jsonify({"success": False, "error": "Missing required fields"}), 400
        
        db_session = SessionLocal()
        try:
            # Check if target entry exists for this year, create if not
            target_entry = db_session.query(models.Report_TvA).filter(
                models.Report_TvA.year == year
            ).first()
            
            if not target_entry:
                # Create new entry with all zeros
                target_entry = models.Report_TvA(
                    year=year,
                    classic=0,
                    bronze=0,
                    silver=0,
                    gold=0,
                    platinum=0,
                    safe_card=0,
                    senior=0,
                    senior_plus=0
                )
                db_session.add(target_entry)
                db_session.flush()  # Get the ID without committing
                print(f"✅ Auto-created target row for year {year}")
            
            # Update the specific category
            if category == "Classic":
                target_entry.classic = target_count
            elif category == "Bronze":
                target_entry.bronze = target_count
            elif category == "Silver":
                target_entry.silver = target_count
            elif category == "Gold":
                target_entry.gold = target_count
            elif category == "Platinum":
                target_entry.platinum = target_count
            elif category == "Safe Card":
                target_entry.safe_card = target_count
            elif category == "Senior":
                target_entry.senior = target_count
            elif category == "Senior+":
                target_entry.senior_plus = target_count
            
            db_session.commit()
            
            # Log the action
            db_conn.POST_action_log(
                current_user.username, 
                current_user.user_level, 
                'Save Target Data', 
                f'Saved {category} target: {target_count} for {year}', 
                current_user.account_id
            )
            
            return jsonify({"success": True, "message": "Target data saved successfully"})
            
        except Exception as e:
            db_session.rollback()
            print(f"❌ Database error: {e}")
            return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500
        finally:
            db_session.close()
        
    except Exception as e:
        print(f"Error saving target vs actual data: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@server.route('/api/save_report/budget_expenses', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def save_budget_expenses():
    """Save budget vs expenses data"""
    try:
        data = request.get_json()
        print("Saving budget expenses data:", data)
        
        # Extract data from request
        account_code = data.get('accountCode')
        account_name = data.get('accountName')
        budget_amount = data.get('budgetAmount')
        expense_month = data.get('expenseMonth')
        expense_amount = data.get('expenseAmount')
        budget_year = data.get('budgetYear')
        
        # Validate required fields
        if not all([account_code, account_name, budget_amount, expense_month, expense_amount, budget_year]):
            return jsonify({"success": False, "error": "Missing required fields"}), 400
        
        # Here you would save to your database
        # For now, just log and return success
        print(f"💾 Would save: Account={account_code}, Budget={budget_amount}, Month={expense_month}, Expense={expense_amount}, Year={budget_year}")
        
        # Log the action
        db_conn.POST_action_log(
            current_user.username, 
            current_user.user_level, 
            'Save Budget Data', 
            f'Saved budget data: {account_code} - {account_name}', 
            current_user.account_id
        )
        
        return jsonify({"success": True, "message": "Budget data saved successfully"})
        
    except Exception as e:
        print(f"Error saving budget expenses data: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@server.route('/api/inventory/add_stock', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def add_inventory_stock():
    try:
        # ✅ FIX: Save user data to variables first
        user_id = current_user.account_id
        username = current_user.username
        user_level = current_user.user_level
        
        data = request.get_json()
        
        # Extract form data
        category = data.get('category')
        prefix = data.get('prefix')
        start_num = data.get('start_num')
        count = data.get('count')
        
        # Validate required fields
        if not all([category, prefix, start_num, count]):
            db_conn.POST_action_log(username, user_level, 'Add Inventory Failed', 'Missing required fields', user_id)
            return jsonify({
                "success": False, 
                "error": "All fields are required"
            }), 400
        
        # Convert to integers
        try:
            start_num = int(start_num)
            count = int(count)
        except ValueError:
            return jsonify({
                "success": False, 
                "error": "Invalid number format"
            }), 400
        
        # Validate count
        if count <= 0:
            return jsonify({
                "success": False, 
                "error": "Count must be greater than 0"
            }), 400
        
        # Validate count isn't too large (performance protection)
        if count > 10000:
            return jsonify({
                "success": False,
                "error": "Cannot add more than 10,000 IDs at once"
            }), 400
        
        # Call the database function
        result = db_conn.add_inventory_ids(category, prefix, start_num, count)
        
        if result["success"]:
            db_conn.POST_action_log(username, user_level, 'Add Inventory', f'Added {result["added_count"]} {category} IDs starting from {prefix}{start_num}', user_id)
            
            response_data = {
                "success": True,
                "message": f"Successfully added {result['added_count']} ID(s) for {category} category",
                "details": result
            }
            
            # Add warning if there were duplicates
            if result['duplicate_count'] > 0:
                response_data["warning"] = f"Skipped {result['duplicate_count']} duplicate ID(s)"
                
            return jsonify(response_data)
        else:
            db_conn.POST_action_log(username, user_level, 'Add Inventory Failed', f'Failed to add {category} IDs: {result.get("error")}', user_id)
            return jsonify({
                "success": False,
                "error": result.get("error", "Failed to add IDs to inventory")
            }), 500
            
    except Exception as e:
        print(f"Error adding inventory stock: {e}")
        # ✅ FIX: Use the saved variables here too
        db_conn.POST_action_log(username, user_level, 'Add Inventory Error', f'Error: {str(e)}', user_id)
        return jsonify({
            "success": False,
            "error": "Internal server error"
        }), 500
    
# ============================================================
# 🔹 ASSIGN ID TO MEMBER - USING allocated_to FOR MEMBER NAMES
# ============================================================

@server.route('/api/inventory/assign_id', methods=['POST'])
@login_required
def assign_id_to_member():
    try:
        data = request.get_json()
        print(f"🔍 Assign ID Request Data: {data}")  # Debug log
        
        # Validate required fields
        required_fields = ['category', 'id_number', 'member_name']
        for field in required_fields:
            if field not in data or not data[field]:
                print(f"❌ Missing field: {field}")
                return jsonify({
                    'success': False,
                    'error': f'Missing required field: {field}'
                }), 400
        
        category = data['category']
        id_number = data['id_number']
        member_name = data['member_name']
        
        print(f"🔍 Processing assignment: {id_number} to {member_name} in category {category}")
        
        # ✅ FIXED: More flexible MAAB number format validation
        import re
        # Allow formats like: PC1206077, PC0004323, etc.
        if not re.match(r'^(PC|PB|PS|PG|PP|PEP|S|SP)\d+$', id_number):
            print(f"❌ Invalid MAAB format: {id_number}")
            return jsonify({
                'success': False,
                'error': f'Invalid MAAB number format: {id_number}. Expected format: Prefix + Numbers (e.g., PC1206077)'
            }), 400
        
        # Check if ID exists and is available
        db_session = SessionLocal()
        
        try:
            # Check if the ID exists and is available
            inventory_item = db_session.query(models.Inventory).filter(
                models.Inventory.maab_no == id_number,
                models.Inventory.maab_category == category
            ).first()
            
            print(f"🔍 Found inventory item: {inventory_item}")
            if inventory_item:
                print(f"🔍 Inventory item details - used: {inventory_item.used}, allocated_to: {inventory_item.allocated_to}")
            
            if not inventory_item:
                return jsonify({
                    'success': False,
                    'error': f'ID {id_number} not found in category {category}'
                }), 400
            
            # Check if already used (used = 1 means used, 0 means available)
            if inventory_item.used == 1:
                return jsonify({
                    'success': False,
                    'error': f'ID {id_number} is already assigned to {inventory_item.allocated_to}'
                }), 400
            
            # Update the inventory item - use allocated_to for member name
            inventory_item.used = 1  # Set to used
            inventory_item.allocated_to = member_name  # Store member name here
            inventory_item.updated_at = datetime.now()
            
            db_session.commit()
            
            print(f"✅ Successfully assigned {id_number} to {member_name}")
            
            # Log the action
            db_conn.POST_action_log(
                current_user.username, 
                current_user.user_level, 
                'Assign ID', 
                f'Assigned {id_number} to {member_name}', 
                current_user.account_id
            )
            
            return jsonify({
                'success': True,
                'message': f'ID {id_number} successfully assigned to {member_name}'
            })
            
        except Exception as e:
            db_session.rollback()
            print(f"❌ Database error in assign_id_to_member: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': f'Database error: {str(e)}'
            }), 500
            
        finally:
            db_session.close()
            
    except Exception as e:
        print(f"❌ Error in assign_id_to_member route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Internal server error: {str(e)}'
        }), 500

# ============================================================
# 🔹 GET AVAILABLE IDs FOR CATEGORY
# ============================================================

@server.route('/api/inventory/available_ids/<category>', methods=['GET'])
@login_required
def get_available_ids(category):
    try:
        count = request.args.get('count', 1, type=int)
        
        print(f"🔍 Getting {count} available IDs for category: {category}")
        
        db_session = SessionLocal()
        
        try:
            # Get available IDs for the category - used = 0 means available
            available_ids = db_session.query(models.Inventory).filter(
                models.Inventory.maab_category == category,
                models.Inventory.used == 0  # 0 = available, 1 = used
            ).limit(count).all()
            
            ids_list = [item.maab_no for item in available_ids]
            
            print(f"✅ Found {len(ids_list)} available IDs: {ids_list}")
            
            return jsonify({
                'success': True,
                'available_ids': ids_list,
                'count': len(ids_list)
            })
            
        except Exception as e:
            print(f"❌ Database error in get_available_ids: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': 'Database error occurred'
            }), 500
            
        finally:
            db_session.close()
            
    except Exception as e:
        print(f"❌ Error in get_available_ids route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500
    
# ============================================================
# 🔹 GET ASSIGNED IDs (For Verification)
# ============================================================

@server.route('/api/inventory/assigned_ids', methods=['GET'])
@login_required
def get_assigned_ids():
    try:
        db_session = SessionLocal()
        
        try:
            # Get assigned IDs (used = 1)
            assigned_ids = db_session.query(models.Inventory).filter(
                models.Inventory.used == 1
            ).all()
            
            result = []
            for item in assigned_ids:
                result.append({
                    'maab_no': item.maab_no,
                    'category': item.maab_category,
                    'allocated_to': item.allocated_to,
                    'remarks': item.remarks
                })
            
            return jsonify({
                'success': True,
                'assigned_ids': result,
                'count': len(result)
            })
            
        except Exception as e:
            print(f"❌ Database error in get_assigned_ids: {e}")
            return jsonify({
                'success': False,
                'error': 'Database error occurred'
            }), 500
            
        finally:
            db_session.close()
            
    except Exception as e:
        print(f"❌ Error in get_assigned_ids route: {e}")
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500
    
# ============================================================
# 🔹 GET ALL AVAILABLE IDs (For MAAB No Dropdown)
# ============================================================

@server.route('/api/inventory/available_ids', methods=['GET'])
@login_required
def get_all_available_ids():
    try:
        category = request.args.get('category')
        
        print(f"🔍 Getting available IDs for category: {category}")
        
        db_session = SessionLocal()
        
        try:
            if category:
                # Get available IDs for specific category - used = 0 means available
                available_ids = db_session.query(models.Inventory).filter(
                    models.Inventory.maab_category == category,
                    models.Inventory.used == 0  # 0 = available, 1 = used
                ).order_by(models.Inventory.maab_no).all()
                
                ids_list = [item.maab_no for item in available_ids]
                
                result = {
                    category: ids_list
                }
                
            else:
                # Get all available IDs grouped by category
                available_ids = db_session.query(models.Inventory).filter(
                    models.Inventory.used == 0
                ).order_by(models.Inventory.maab_category, models.Inventory.maab_no).all()
                
                result = {}
                for item in available_ids:
                    if item.maab_category not in result:
                        result[item.maab_category] = []
                    result[item.maab_category].append(item.maab_no)
            
            print(f"✅ Found available IDs: { {k: len(v) for k, v in result.items()} }")
            
            return jsonify({
                'success': True,
                'available_ids': result
            })
            
        except Exception as e:
            print(f"❌ Database error in get_all_available_ids: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': 'Database error occurred'
            }), 500
            
        finally:
            db_session.close()
            
    except Exception as e:
        print(f"❌ Error in get_all_available_ids route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500

# ============================================================
# 🔹 REPAIR INVENTORY SYNC
# ============================================================

@server.route('/api/inventory/repair_sync', methods=['POST'])
@login_required
def repair_inventory_sync():
    """Repair inventory sync by marking MAAB numbers as used if they exist in entries"""
    try:
        db_session = SessionLocal()
        
        try:
            # Get all MAAB numbers that are actually used in entries
            cursor = db_session.execute(text("""
                SELECT DISTINCT maab_no 
                FROM entry_contents 
                WHERE maab_no IS NOT NULL AND maab_no != ''
            """))
            used_maab_numbers = [row[0] for row in cursor.fetchall()]
            
            print(f"🔍 Found {len(used_maab_numbers)} used MAAB numbers in entries")
            
            # Mark these as used in inventory
            updated_count = 0
            for maab_no in used_maab_numbers:
                inventory_item = db_session.query(models.Inventory).filter(
                    models.Inventory.maab_no == maab_no
                ).first()
                
                if inventory_item and inventory_item.used == 0:
                    inventory_item.used = 1
                    inventory_item.updated_at = datetime.now()
                    updated_count += 1
                    print(f"✅ Fixed: {maab_no} marked as used in inventory")
            
            db_session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Repaired {updated_count} inventory items. Now in sync with entries.',
                'updated_count': updated_count,
                'total_used_in_entries': len(used_maab_numbers)
            })
            
        except Exception as e:
            db_session.rollback()
            print(f"❌ Database error in repair_sync: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': 'Database error occurred'
            }), 500
            
        finally:
            db_session.close()
            
    except Exception as e:
        print(f"❌ Error in repair_sync route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500
    
# ============================================================
# 🔹 TEST INVENTORY ASSIGNMENT
# ============================================================

@server.route('/api/debug/test_assignment', methods=['POST'])
@login_required
def test_assignment():
    """Test route to check if assignment works"""
    try:
        db_session = SessionLocal()
        
        # Get one available ID
        available_item = db_session.query(models.Inventory).filter(
            models.Inventory.used == 0
        ).first()
        
        if not available_item:
            return jsonify({'success': False, 'error': 'No available IDs found'})
        
        test_data = {
            'category': available_item.maab_category,
            'id_number': available_item.maab_no,
            'member_name': 'TEST MEMBER'
        }
        
        return jsonify({
            'success': True,
            'test_data': test_data,
            'message': 'Use this data to test assignment'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    finally:
        db_session.close()

# ============================================================
# 🔹 MARK MAAB NUMBER AS USED IN INVENTORY
# ============================================================

@server.route('/api/inventory/mark_as_used', methods=['POST'])
@login_required
def mark_maab_as_used():
    """Mark a MAAB number as used in inventory"""
    try:
        data = request.get_json()
        maab_no = data.get('maab_no')
        
        if not maab_no:
            return jsonify({'success': False, 'error': 'MAAB number is required'})
        
        print(f"🔍 Marking MAAB number as used: {maab_no}")
        
        db_session = SessionLocal()
        
        try:
            # Find the inventory item
            inventory_item = db_session.query(models.Inventory).filter(
                models.Inventory.maab_no == maab_no
            ).first()
            
            if not inventory_item:
                return jsonify({'success': False, 'error': f'MAAB number {maab_no} not found in inventory'})
            
            if inventory_item.used == 1:
                return jsonify({'success': False, 'error': f'MAAB number {maab_no} is already marked as used'})
            
            # Mark as used
            inventory_item.used = 1
            inventory_item.updated_at = datetime.now()
            
            db_session.commit()
            
            print(f"✅ Successfully marked {maab_no} as used in inventory")
            
            return jsonify({
                'success': True,
                'message': f'MAAB number {maab_no} marked as used'
            })
            
        except Exception as e:
            db_session.rollback()
            print(f"❌ Database error in mark_maab_as_used: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': 'Database error occurred'
            }), 500
            
        finally:
            db_session.close()
            
    except Exception as e:
        print(f"❌ Error in mark_maab_as_used route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500
    
@server.route('/api/add_to_dispatch', methods=['PATCH'])
@login_required
@roles_required('admin', 'superadmin')
def add_to_dispatch():
    data = request.get_json() or {}
    record_ids = data.get('record_ids', [])
    
    try:
        print(f"📦 API: Adding records to dispatch: {record_ids}")
        result = db_conn.add_to_dispatch(record_ids)
        print(f'📦 API: add_to_dispatch result: {result}')
        
        if result is not None:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Add to Dispatch', f'Added {result} entries to dispatch', current_user.account_id)

            return jsonify({"success": True, "added_to_dispatch_count": result})
        else:
            return jsonify({"success": False, "error": "No entries to add to dispatch"}), 500
            
    except Exception as e:
        print(f"❌ API Error in add_to_dispatch route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
    
@server.route('/api/cancel_dispatch', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def cancel_dispatch():
    """Cancel the current active dispatch by setting status to 'disregard'"""
    db_session = SessionLocal()
    try:
        print("🎯 Starting cancel dispatch process...")
        
        # Get the active dispatch using THE SAME session
        active_dispatch = db_session.query(models.Dispatch).filter(
            models.Dispatch.dispatch_status == 'current'
        ).order_by(models.Dispatch.dispatch_id.desc()).first()
        
        if not active_dispatch:
            print("❌ No active dispatch found")
            return jsonify({"success": False, "error": "No active dispatch found"}), 400

        dispatch_id = active_dispatch.dispatch_id
        original_status = active_dispatch.dispatch_status
        
        print(f"🎯 Canceling dispatch: {dispatch_id}")
        print(f"📊 Current status: {original_status}")

        entries_updated = 0
        records_updated = 0

        # 1. Remove dispatch_id from all entries in this dispatch
        try:
            entries_result = db_session.query(models.Entries).filter(
                models.Entries.dispatch_id == dispatch_id
            ).update({
                models.Entries.dispatch_id: None
            }, synchronize_session=False)
            entries_updated = entries_result if entries_result else 0
            print(f"✅ Entries updated: {entries_updated}")
        except Exception as e:
            print(f"⚠️ No entries to update or error updating entries: {e}")
            entries_updated = 0

        # 2. Remove dispatch_id from all records in this dispatch
        try:
            records_result = db_session.query(models.Records).filter(
                models.Records.dispatch_id == dispatch_id
            ).update({
                models.Records.dispatch_id: None
            }, synchronize_session=False)
            records_updated = records_result if records_result else 0
            print(f"✅ Records updated: {records_updated}")
        except Exception as e:
            print(f"⚠️ No records to update or error updating records: {e}")
            records_updated = 0

        # 3. Update dispatch status to 'disregard'
        active_dispatch.dispatch_status = 'disregard'
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M')
        if active_dispatch.dispatch_remarks:
            active_dispatch.dispatch_remarks = f"{active_dispatch.dispatch_remarks} - Cancelled by {current_user.username} on {current_time}"
        else:
            active_dispatch.dispatch_remarks = f"Cancelled by {current_user.username} on {current_time}"

        print(f"🔍 Before commit - Status: {active_dispatch.dispatch_status}")
        
        # COMMIT THE CHANGES
        db_session.commit()
        
        # Get the final status BEFORE closing the session
        final_status = active_dispatch.dispatch_status
        
        print(f"🔍 After commit - Status: {final_status}")
        print(f"✅ Dispatch {dispatch_id} cancelled successfully")

        # Log the action
        db_conn.POST_action_log(
            current_user.username, 
            current_user.user_level, 
            'Cancel Dispatch', 
            f'Cancelled dispatch {dispatch_id} (entries: {entries_updated}, records: {records_updated})', 
            current_user.account_id
        )

        return jsonify({
            "success": True,
            "message": "Dispatch cancelled successfully",
            "entries_updated": entries_updated,
            "records_updated": records_updated,
            "new_status": final_status,
            "dispatch_id": dispatch_id
        })

    except Exception as e:
        db_session.rollback()
        print(f"❌ Error cancelling dispatch: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        db_session.close()

@server.route('/api/members/get_next_record_id')
@login_required
def get_next_record_id():
    """Get the next available record ID without reserving it"""
    db_session = SessionLocal()
    try:
        # Get the maximum record_id currently in use
        max_id = db_session.query(func.max(models.Records.record_id)).scalar()
        next_id = (max_id or 0) + 1
        print(f"Next available record ID: {next_id}")
        
        return jsonify({"next_record_id": next_id})
    except Exception as e:
        print(f"Error getting next record ID: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"next_record_id": 0, "error": str(e)}), 500
    finally:
        db_session.close()
                        

# FIXED EXPORT DISPATCH ROUTE - ONLY ONE VERSION
@server.route('/api/export_dispatch', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def export_dispatch():
    try:
        data = request.get_json()
        if not data:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Export Data Failed', 'No JSON data provided', current_user.account_id)
            return jsonify({"success": False, "error": "No JSON data provided"}), 400
            
        selected_rows = data.get('selected_rows', [])
        
        if not selected_rows:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Export Data Failed', 'No rows selected', current_user.account_id)
            return jsonify({"success": False, "error": "No rows selected"}), 400

        # Create DataFrame from selected rows
        df_data = []
        for i, row in enumerate(selected_rows, 1):
            df_data.append({
                'No.': i,
                'Name': row.get('name', ''),
                'Category': row.get('category', ''),
                'Effectivity': row.get('effectivity', ''),
                'Birthday': row.get('birthday', ''),
                'Location': row.get('location', '')
            })

        df = pd.DataFrame(df_data)

        # Define the export folder path
        export_folder = os.path.join(os.path.dirname(__file__), 'exports', 'dispatch_reports')
        
        # Create folder if it doesn't exist
        os.makedirs(export_folder, exist_ok=True)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Dispatch_Report_{timestamp}.xlsx"
        file_path = os.path.join(export_folder, filename)

        # Create Excel writer with styling
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            # Write the data
            df.to_excel(writer, sheet_name='Dispatch Report', index=False, startrow=2)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Dispatch Report']
            
            # Add title
            title_format = workbook.add_format({
                'bold': True,
                'size': 16,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            worksheet.merge_range('A1:F1', 'DISPATCH REPORT', title_format)
            
            # Header format
            header_format = workbook.add_format({
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#D3D3D3',
                'border': 1
            })
            
            # Apply header format
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(2, col_num, value, header_format)
            
            # Set column widths
            worksheet.set_column('A:A', 5)   # No.
            worksheet.set_column('B:B', 30)  # Name
            worksheet.set_column('C:C', 15)  # Category
            worksheet.set_column('D:D', 12)  # Effectivity
            worksheet.set_column('E:E', 12)  # Birthday
            worksheet.set_column('F:F', 25)  # Location
            
            # Center the number column
            number_format = workbook.add_format({'align': 'center'})
            worksheet.set_column('A:A', 5, number_format)

        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Export Data', f'Exported dispatch report with {len(selected_rows)} entries', current_user.account_id)
        return jsonify({
            "success": True, 
            "message": "File saved successfully",
            "filename": filename,
            "file_path": file_path
        })

    except Exception as e:
        print(f"Export error: {e}")
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Export Data Error', f'Error: {str(e)}', current_user.account_id)
        return jsonify({"success": False, "error": str(e)}), 500


def generate_dispatch_excel_file(dispatch_id, transmitted_count):
    """Generate multi-sheet Excel file for the dispatched entries"""
    try:
        print(f"🎯 Generating multi-sheet Excel for dispatch {dispatch_id}")
        
        db_session = SessionLocal()
        
        # Get dispatch details
        dispatch = db_session.query(models.Dispatch).filter_by(dispatch_id=dispatch_id).first()
        if not dispatch:
            print(f"❌ Dispatch {dispatch_id} not found")
            return jsonify({"success": False, "error": "Dispatch not found"}), 404

        print(f"✅ Found dispatch: {dispatch.dispatch_id}")

        # Get entries that were declared today (regardless of dispatch_id)
        current_date = datetime.now().date()
        entries_data = (
            db_session.query(
                models.Entries.entry_id,
                models.Entries.maab_category,
                models.Entries.maab_no,
                models.Members.first_name,
                models.Members.middle_name,
                models.Members.last_name,
                models.Members.suffix,
                models.Members.birth_date,
                models.Members.age,
                models.Members.sex,
                models.Members.contact_no,
                models.Members.email,
                models.Members.address,
                models.Records.effectivity_date,
                models.Records.location_particular,
                models.Records.location_category,
                models.Records.municipality,
                models.Records.district,
                models.Records.origin
            )
            .join(models.Members, models.Entries.member_id == models.Members.member_id)
            .join(models.Records, models.Entries.record_id == models.Records.record_id)
            .filter(
                models.Entries.declaration_date == current_date,
                models.Records.declaration_date == current_date
            )
            .all()
        )

        print(f"✅ Found {len(entries_data)} entries declared today")

        # If no entries found, create a simple success response
        if not entries_data:
            print("⚠️ No entries found declared today, creating simple success response")
            return jsonify({
                "success": True,
                "message": f"Dispatch transmitted successfully! {transmitted_count} entries processed.",
                "transmitted_count": transmitted_count,
                "excel_generated": False
            })

        # Create Excel file in memory
        output = BytesIO()
        
        try:
            # Use openpyxl engine for better multi-sheet support
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # Define custom colors
                header_fill = openpyxl.styles.PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")  # Purple-gray
                total_row_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray
                red_font = openpyxl.styles.Font(color="FF0000", bold=True)  # Red font for numbers
                bold_font = openpyxl.styles.Font(bold=True)
                
                # ==================== LISTING SHEET (Summary) ====================
                print("📊 Creating Listing sheet...")
                
                # Group by location_particular and category
                summary_data = []
                location_groups = {}
                
                for entry in entries_data:
                    location = entry.location_particular or 'Unknown'
                    category = entry.maab_category or 'Unknown'
                    
                    if location not in location_groups:
                        location_groups[location] = {
                            'location': location,
                            'effectivity_date': entry.effectivity_date,
                            'categories': {}
                        }
                    
                    if category not in location_groups[location]['categories']:
                        location_groups[location]['categories'][category] = 0
                    
                    location_groups[location]['categories'][category] += 1
                
                # Convert to list for DataFrame
                categories_list = ['Classic', 'Bronze', 'Silver', 'Gold', 'Platinum', 'Safe Card', 'Senior', 'Senior Plus']
                for i, (location, data) in enumerate(location_groups.items(), 1):
                    row = {
                        'NO.': i,
                        'SCHOOLS/COMPANY': location,
                        'EFFECTIVITY DATE': data['effectivity_date'].strftime('%Y-%m-%d') if data['effectivity_date'] else ''
                    }
                    
                    # Add counts for each category
                    for category in categories_list:
                        row[category.upper()] = data['categories'].get(category, 0)
                    
                    summary_data.append(row)
                
                # Add empty row before totals
                empty_row = {'NO.': '', 'SCHOOLS/COMPANY': '', 'EFFECTIVITY DATE': ''}
                for category in categories_list:
                    empty_row[category.upper()] = ''
                summary_data.append(empty_row)
                
                # Add totals row
                if summary_data:
                    totals_row = {'NO.': '', 'SCHOOLS/COMPANY': 'TOTAL', 'EFFECTIVITY DATE': ''}
                    for category in categories_list:
                        totals_row[category.upper()] = sum(row[category.upper()] for row in summary_data if row['SCHOOLS/COMPANY'] != 'TOTAL' and row['SCHOOLS/COMPANY'] != '')
                    summary_data.append(totals_row)
                
                # Create summary DataFrame
                summary_df = pd.DataFrame(summary_data)
                
                # Write summary sheet
                summary_df.to_excel(writer, sheet_name='Listing', index=False, startrow=2)
                
                # Style summary sheet
                workbook = writer.book
                summary_sheet = writer.sheets['Listing']
                
                # Add title
                summary_sheet.merge_cells('A1:K1')
                title_cell = summary_sheet['A1']
                title_cell.value = f'DISPATCH REPORT - CHAPTER - {dispatch.dispatch_cutoff}'
                title_cell.font = openpyxl.styles.Font(size=16, bold=True)
                title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Style headers and NO. column
                for col in range(1, len(summary_df.columns) + 1):
                    cell = summary_sheet.cell(row=3, column=col)
                    cell.fill = header_fill
                    cell.font = bold_font
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Style NO. column for all data rows
                data_row_count = len(location_groups)  # Number of actual data rows (excluding empty and total)
                for row_num in range(4, 4 + data_row_count):  # Start from row 4 (after header)
                    no_cell = summary_sheet.cell(row=row_num, column=1)  # Column A (NO.)
                    no_cell.fill = header_fill
                
                # Style totals row
                if summary_data:
                    total_row_num = 4 + data_row_count + 1  # +1 for the empty row
                    for col in range(1, len(summary_df.columns) + 1):
                        cell = summary_sheet.cell(row=total_row_num, column=col)
                        cell.fill = total_row_fill
                        cell.font = bold_font
                        
                        # Make numbers red in totals row (columns D-K)
                        if col >= 4:  # Columns D onwards (category counts)
                            cell.font = red_font
                
                # Adjust column widths for summary sheet
                summary_widths = {
                    'A': 8,   # NO.
                    'B': 35,  # SCHOOLS/COMPANY
                    'C': 15,  # EFFECTIVITY DATE
                    'D': 10,  # CLASSIC
                    'E': 10,  # BRONZE
                    'F': 10,  # SILVER
                    'G': 10,  # GOLD
                    'H': 10,  # PLATINUM
                    'I': 12,  # SAFE CARD
                    'J': 10,  # SENIOR
                    'K': 12   # SENIOR PLUS
                }
                
                for col_letter, width in summary_widths.items():
                    summary_sheet.column_dimensions[col_letter].width = width
                
                # ==================== INDIVIDUAL CATEGORY SHEETS ====================
                categories = ['Classic', 'Bronze', 'Silver', 'Gold', 'Platinum', 'Safe Card', 'Senior', 'Senior Plus']
                
                for category in categories:
                    print(f"📝 Creating {category} sheet...")
                    
                    # Filter entries for this category
                    category_entries = [e for e in entries_data if e.maab_category == category]
                    
                    if not category_entries:
                        print(f"⚠️ No entries found for {category}, creating empty sheet")
                        # Create empty DataFrame with correct columns
                        empty_data = []
                        empty_df = pd.DataFrame(empty_data, columns=['NO.', 'NAME', 'PRC ID #', 'EFFECTIVITY', 'BIRTHDAY', 'ADDRESS'])
                        empty_df.to_excel(writer, sheet_name=category, index=False, startrow=2)
                    else:
                        # Prepare data for this category
                        category_data = []
                        for i, entry in enumerate(category_entries, 1):
                            full_name = f"{entry.first_name or ''} {entry.middle_name or ''} {entry.last_name or ''} {entry.suffix or ''}".strip()
                            category_data.append({
                                'NO.': i,
                                'NAME': full_name,
                                'PRC ID #': entry.maab_no or '',
                                'EFFECTIVITY': entry.effectivity_date.strftime('%Y-%m-%d') if entry.effectivity_date else '',
                                'BIRTHDAY': entry.birth_date.strftime('%Y-%m-%d') if entry.birth_date else '',
                                'ADDRESS': entry.address or ''
                            })
                        
                        # Create category DataFrame
                        category_df = pd.DataFrame(category_data)
                        category_df.to_excel(writer, sheet_name=category, index=False, startrow=2)
                    
                    # Style category sheet
                    category_sheet = writer.sheets[category]
                    
                    # Add title
                    category_sheet.merge_cells('A1:F1')
                    title_cell = category_sheet['A1']
                    title_cell.value = f'DISPATCH REPORT - CHAPTER - {dispatch.dispatch_cutoff}'
                    title_cell.font = openpyxl.styles.Font(size=16, bold=True)
                    title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    
                    # Style headers and NO. column for category sheets
                    for col in range(1, 7):  # A-F
                        cell = category_sheet.cell(row=3, column=col)
                        cell.fill = header_fill
                        cell.font = bold_font
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    
                    # Style NO. column for all data rows in category sheets
                    if category_entries:
                        for row_num in range(4, 4 + len(category_entries)):
                            no_cell = category_sheet.cell(row=row_num, column=1)  # Column A (NO.)
                            no_cell.fill = header_fill
                    
                    # Adjust column widths for category sheets
                    category_widths = {
                        'A': 8,   # NO.
                        'B': 30,  # NAME
                        'C': 15,  # PRC ID #
                        'D': 12,  # EFFECTIVITY
                        'E': 12,  # BIRTHDAY
                        'F': 40   # ADDRESS
                    }
                    
                    for col_letter, width in category_widths.items():
                        category_sheet.column_dimensions[col_letter].width = width

            # Prepare file for download
            output.seek(0)
            
            # Create filename
            filename = f"dispatch_{dispatch.dispatch_origin}_{current_date.strftime('%Y-%m-%d')}.xlsx"
            
            print(f"✅ Multi-sheet Excel file generated successfully: {filename}")
            
            # Return the file
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as excel_error:
            print(f"❌ Excel generation error: {excel_error}")
            # Fallback: Return success without Excel file
            return jsonify({
                "success": True,
                "message": f"Dispatch transmitted successfully! {transmitted_count} entries processed. (Excel generation skipped)",
                "transmitted_count": transmitted_count,
                "excel_generated": False
            })

    except Exception as e:
        print(f"❌ Error in generate_dispatch_excel_file: {e}")
        import traceback
        traceback.print_exc()
        # Fallback: Return success without Excel file
        return jsonify({
            "success": True,
            "message": f"Dispatch transmitted successfully! {transmitted_count} entries processed. (Excel generation failed)",
            "transmitted_count": transmitted_count,
            "excel_generated": False
        })
    finally:
        try:
            db_session.close()
        except:
            pass

def create_fallback_excel_file(dispatch, transmitted_count):
    """Create a simple Excel file when no entries are found"""
    try:
        # Create simple data with just dispatch info
        data = [{
            'Dispatch ID': dispatch.dispatch_id,
            'Dispatch Type': dispatch.dispatch_type,
            'Dispatch Origin': dispatch.dispatch_origin,
            'Dispatch Year': dispatch.dispatch_year,
            'Dispatch Cutoff': dispatch.dispatch_cutoff.strftime('%Y-%m-%d') if dispatch.dispatch_cutoff else 'N/A',
            'Date Dispatched': dispatch.date_dispatched.strftime('%Y-%m-%d') if dispatch.date_dispatched else 'N/A',
            'Total Entries': transmitted_count,
            'Late Declare': dispatch.late_declare,
            'Remarks': dispatch.dispatch_remarks or 'No remarks'
        }]

        df = pd.DataFrame(data)

        # Create Excel file in memory
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Dispatch Summary', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Dispatch Summary']
            
            # Add title
            worksheet.merge_cells('A1:I1')
            title_cell = worksheet['A1']
            title_cell.value = f'DISPATCH SUMMARY - {dispatch.dispatch_origin}'
            title_cell.font = openpyxl.styles.Font(size=16, bold=True)
            title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        output.seek(0)
        
        filename = f"dispatch_summary_{dispatch.dispatch_origin}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error creating fallback Excel: {e}")
        # If even the fallback fails, return success without file
        return jsonify({
            "success": True, 
            "message": f"Dispatch transmitted successfully! {transmitted_count} entries processed.",
            "transmitted_count": transmitted_count
        })

@server.route('/api/remove_from_dispatch', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def remove_from_dispatch():
    print("🎯🎯🎯 REMOVE FROM DISPATCH ROUTE HIT! 🎯🎯🎯")
    
    try:
        data = request.get_json()
        print(f"Received data: {data}")
        
        if not data:
            return jsonify({"success": False, "error": "No data provided"}), 400
            
        selected_rows = data.get('selected_rows', [])
        print(f"Selected rows to remove: {selected_rows}")
        
        if not selected_rows:
            return jsonify({"success": False, "error": "No entries selected"}), 400
        
        # Call database function to remove entries from dispatch
        result = db_conn.remove_entries_from_dispatch(selected_rows)
        
        print(f"📦 Remove from dispatch result: {result}")
        
        if result["success"]:
            return jsonify({
                "success": True, 
                "removed_count": result["removed_count"],
                "message": f"Successfully removed {result['removed_count']} entries from dispatch"
            })
        else:
            return jsonify({"success": False, "error": result["error"]}), 500
            
    except Exception as e:
        print(f"❌ Remove from dispatch error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": "Internal server error"}), 500
        
@server.route('/api/transmit_dispatch', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def transmit_dispatch():
    print("🎯🎯🎯 TRANSMIT DISPATCH ROUTE HIT! 🎯🎯🎯")
    
    try:
        data = request.get_json() or {}
        print(f"Received data: {data}")
        
        # Get the active dispatch
        active_dispatch = db_conn.get_current_active_dispatch()
        if not active_dispatch:
            print("❌ No active dispatch found")
            return jsonify({"success": False, "error": "No active dispatch found"}), 400

        print(f"✅ Active dispatch found: {active_dispatch.dispatch_id}")

        # Call the transmission function
        result = db_conn.transmit_dispatch_entries(active_dispatch.dispatch_id, current_user.account_id)
        
        print(f"📦 Transmission result: {result}")
        
        if result["success"]:
            try:
                # Try to generate Excel file
                return generate_dispatch_excel_file(active_dispatch.dispatch_id, result['transmitted_count'])
            except Exception as excel_error:
                print(f"❌ Excel generation failed, returning JSON success: {excel_error}")
                # Fallback to JSON response if Excel fails
                return jsonify({
                    "success": True,
                    "message": f"Dispatch transmitted successfully! {result['transmitted_count']} entries processed.",
                    "transmitted_count": result['transmitted_count'],
                    "dispatch_id": active_dispatch.dispatch_id
                })
        else:
            return jsonify({"success": False, "error": result["error"]}), 500
            
    except Exception as e:
        print(f"❌ Transmit dispatch error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": "Internal server error"}), 500
    
@server.route('/api/debug_dispatch_state', methods=['GET'])
@login_required
@roles_required('admin', 'superadmin')
def debug_dispatch_state():
    """Debug endpoint to check dispatch state"""
    try:
        db_session = SessionLocal()
        
        # Get active dispatch
        active_dispatch = db_conn.get_current_active_dispatch()
        
        # Get entries in dispatch
        entries_in_dispatch = []
        if active_dispatch:
            entries_in_dispatch = db_session.query(models.Entries).filter_by(dispatch_id=active_dispatch.dispatch_id).all()
        
        result = {
            "active_dispatch": {
                "id": active_dispatch.dispatch_id if active_dispatch else None,
                "status": active_dispatch.dispatch_status if active_dispatch else None,
                "type": active_dispatch.dispatch_type if active_dispatch else None,
                "origin": active_dispatch.dispatch_origin if active_dispatch else None
            } if active_dispatch else None,
            "entries_in_dispatch": len(entries_in_dispatch),
            "entries_details": [
                {
                    "entry_id": entry.entry_id,
                    "record_id": entry.record_id,
                    "dispatch_id": entry.dispatch_id,
                    "dispatch_ready": entry.dispatch_ready,
                    "declared": entry.declared,
                    "maab_no": entry.maab_no
                }
                for entry in entries_in_dispatch
            ]
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"error": str(e)})
    finally:
        db_session.close()

@server.route('/debug/routes')
def debug_routes():
    """Debug endpoint to show all registered routes"""
    routes = []
    for rule in server.url_map.iter_rules():
        routes.append({
            'endpoint': rule.endpoint,
            'methods': list(rule.methods),
            'path': str(rule)
        })
    return jsonify(sorted(routes, key=lambda x: x['path']))

@server.route('/api/test_transmit', methods=['GET'])
def test_transmit():
    return jsonify({"message": "Transmit endpoint is accessible", "status": "success"})

# TODO check if a record exist for Online Registration on the same day. if yes put the entry on that record
@server.route('/api/member_register', methods=['POST'])
def member_register():
    data = request.get_json()
    print(data)
    if data:
        fname = data.get('fname').upper()
        mname = data.get('mname').upper()
        lname = data.get('lname').upper()
        suffix = data.get('suffix')
        
        birthdate_string = data.get('birthdate')
        birthdate = datetime.strptime(birthdate_string, "%Y-%m-%d").date()
        
        age = data.get('age')
        
        sex = data.get('sex')
        if sex == "null" or sex == "":
            sex = None
            
        bloodtype = data.get('bloodtype')
        if bloodtype == "null" or bloodtype == "":
            bloodtype = None
        
        contact = data.get('contact')
        email = data.get('email')
        municipality = data.get('municipality')
        
        address = data.get('address')
        maab_cat = data.get('maab_cat')
        origin = data.get('origin')
        
        result = db_conn.add_entry_content_online(fname, mname, lname, suffix, birthdate, age, sex, bloodtype, contact, email, municipality, address, maab_cat, origin)

        if result:
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": result}), 500
    else:
        return jsonify({"success": False, "error": "No data provided"}), 400
    
@server.route('/api/check_active_dispatch', methods=['GET'])
@login_required
@roles_required('admin', 'superadmin')
def check_active_dispatch():
    try:
        active_dispatch = db_conn.get_current_active_dispatch()
        return jsonify({
            'has_active_dispatch': active_dispatch is not None,
            'dispatch_id': active_dispatch.dispatch_id if active_dispatch else None
        })
    except Exception as e:
        print(f"Error checking active dispatch: {e}")
        return jsonify({'has_active_dispatch': False}), 500
    
@server.route('/api/get_claim_records')
def get_claim_records():
    claim_records = db_conn.get_claim_records()
    claim_records = list(reversed(claim_records))  # reverse the list
    #print('flask_server: claim_records', claim_records)
    return jsonify(claim_records)

    '''
    records = db_conn.get_claim_records()
    claims_list = []

    for row in records:
        record = dict(row._mapping)

        for key in ['same_as_insured', 'picked_up']:
            if key in record:
                record[key] = bool(record[key])

        for key in ['date_filed', 'date_of_loss', 'date_released', 'date_picked_up', 'effectivity_date']:
            if key in record and record[key]:
                record[key] = record[key].strftime('%Y-%m-%d')

        claims_list.append(record)
    return jsonify(claims_list)
    '''

@server.route('/api/get_claim_id', methods=['GET'])
def get_claim_id():
    new_claim_id = db_conn.get_new_claim_id()
    print('flask_server: new_claim_id', new_claim_id)
    return jsonify({"new_claim_id": new_claim_id})

@server.route('/api/add_claim_record', methods=['POST'])
def add_claim_record():
    new_claim_id = db_conn.add_claim_record()
    return jsonify({"success": True, "claim_id": new_claim_id})


@server.route('/api/get_maab_numbers', methods=['GET'])
def get_maab_numbers():
    maab_numbers = db_conn.get_unique_maab_numbers()
    return jsonify({"maab_numbers": maab_numbers})


@server.route('/api/verify_maab_no', methods=['POST'])
def verify_maab_no():
    data = request.get_json()
    maab_no = data.get('maab_no')
    if not maab_no:
        return jsonify({"exists": False, "error": "No MAAB No. provided"}), 400

    result = db_conn.verify_maab_no(maab_no)
    if result is None:
        return jsonify({"exists": False})
    return jsonify(result)


@server.route('/api/save_claim_record', methods=['POST'])
def save_claim_record():
    data = request.get_json()
    if data:
        db_conn.save_claim_record(data)
        return jsonify({"success": True})


@server.route('/api/delete_claim_record', methods=['DELETE'])
def delete_claim_record():
    data = request.get_json()
    claim_id = data.get('claim_id')
    if not claim_id:
        return jsonify({"success": False, "error": "No claim ID provided"}), 400

    success = db_conn.delete_claim_record(claim_id)
    if success:
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "error": "Failed to delete claim record"}), 500


#* -------------------- ACCOUNTS API ROUTES -------------------- *#

@server.route('/api/accounts', methods=['GET'])
@login_required
@roles_required('admin', 'superadmin')
def get_accounts():
    # ACTIVE accounts should only be 'approved' status (EXCLUDE archived)
    active_accounts = db_conn.get_accounts(status=['approved'])
    # PENDING accounts
    pending_accounts = db_conn.get_accounts(status=['pending'])

    return jsonify({
        "active": [acc.to_dict() for acc in active_accounts],
        "pending": [acc.to_dict() for acc in pending_accounts]
    })

@server.route('/api/accounts/<id>/create', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def create_account(id):
    pass


@server.route('/api/accounts/approve', methods=['PATCH'])
@login_required
@roles_required('admin', 'superadmin')
def approve_account():
    data = request.get_json()
    ids = data.get('ids', [])

    for acc_id in ids:
        success = db_conn.approve_account(acc_id)

        if not success:
            return jsonify({"success": False, "error": f"Failed to approve account ID {acc_id}"}), 500
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Approve Accounts', f'Approved {len(ids)} account(s)', current_user.account_id)

    return jsonify({"success": True})

@server.route('/api/accounts/decline', methods=['PATCH'])
@login_required
@roles_required('admin', 'superadmin')
def decline_account():
    data = request.get_json()
    ids = data.get('ids', [])

    for acc_id in ids:
        success = db_conn.decline_account(acc_id)
        
        if not success:
            return jsonify({"success": False, "error": f"Failed to decline account ID {acc_id}"}), 500

    return jsonify({"success": True})

@server.route('/api/accounts/reset', methods=['PATCH'])
@login_required
@roles_required('admin', 'superadmin')
def reset_account():
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        print(f"🔍 RESET PASSWORD API CALLED")
        print(f"📋 Account IDs to reset: {ids}")
        
        if not ids:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Reset Password Failed', 'No account IDs provided', current_user.account_id)
            return jsonify({"success": False, "error": "No account IDs provided"}), 400

        success_count = 0
        failed_ids = []
        
        for acc_id in ids:
            print(f"🔄 Processing account ID: {acc_id}")
            
            success = db_conn.reset_account(acc_id)
            
            if success:
                success_count += 1
                print(f"✅ Successfully reset password for account {acc_id}")
            else:
                failed_ids.append(acc_id)
                print(f"❌ Failed to reset password for account {acc_id}")
                db_conn.POST_action_log(current_user.username, current_user.user_level, 'Reset Password Failed', f'Failed to reset password for account ID {acc_id}', current_user.account_id)
        
        print(f"📊 Reset summary: {success_count} successful, {len(failed_ids)} failed")
        
        if success_count > 0:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Reset Passwords', f'Reset passwords for {success_count} account(s)', current_user.account_id)
            response = {
                "success": True, 
                "reset_count": success_count,
                "message": f"Passwords reset for {success_count} account(s)!"
            }
            if failed_ids:

                response["warning"] = f"Failed to reset {len(failed_ids)} account(s)"
            return jsonify(response)
        else:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Reset Password Failed', f'Failed to reset passwords for all {len(ids)} account(s)', current_user.account_id)
            return jsonify({
                "success": False, 
                "error": f"Failed to reset passwords for all {len(ids)} account(s)"
            }), 500
            
    except Exception as e:
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Reset Password Error', f'Error: {str(e)}', current_user.account_id)
        print(f"❌ ERROR in reset_account route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False, 
            "error": f"Internal server error: {str(e)}"
        }), 500
    
@server.route('/api/accounts/update_userlvl', methods=['PATCH'])
@login_required
@roles_required('admin', 'superadmin')
def update_userlvl():
    try:
        data = request.get_json()
        id = data.get('id')
        user_level = data.get('user_level')
    
        success = db_conn.update_userlvl(id, user_level)

        if success:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Update User Level', f'Updated user level to {user_level} for account {id}', current_user.account_id)
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "Failed to update user level"}), 500
        
    except Exception as e:
        print(f"Error updating user level: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@server.route('/api/accounts/update_ofc', methods=['PATCH'])
@login_required
@roles_required('admin', 'superadmin')
def update_ofc():
    try:
        data = request.get_json()
        id = data.get('id')
        location = data.get('office_location')
    
        success = db_conn.update_ofc(id, location)

        if success:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Update Office Location', f'Updated office location to {location} for account {id}', current_user.account_id)
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "Failed to update office location"}), 500
        
    except Exception as e:
        print(f"Error updating office location: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
# UPDATE YOUR FLASK ROUTES

@server.route('/create_acc_submit', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def create_account_submit():
    
    def username_exists(username):
        """Check if username already exists in database using SQLAlchemy"""
        db_session = db_conn.SessionLocal()
        try:
            existing_user = db_session.query(db_conn.models.Accounts).filter(
                db_conn.models.Accounts.username == username
            ).first()
            return existing_user is not None
        finally:
            db_session.close()

    def email_exists(email):
        """Check if email already exists in database using SQLAlchemy"""
        if not email:  # Email is optional
            return False
            
        db_session = db_conn.SessionLocal()
        try:
            existing_email = db_session.query(db_conn.models.Accounts).filter(
                db_conn.models.Accounts.email == email
            ).first()
            return existing_email is not None
        finally:
            db_session.close()

    try:
        data = request.get_json()
        print("=== CREATE ACCOUNT DEBUG ===")
        print("Received data:", data)
        
        # Validate required fields - UPDATED: removed password from required fields
        required_fields = ['username', 'fname', 'lname', 'contact_no', 'branch']
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            print(f"Missing fields: {missing_fields}")
            # ADD FAILED LOG:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Account Failed', f'Missing required fields: {", ".join(missing_fields)}', current_user.account_id)
            return jsonify({
                "success": False, 
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400

        # Validate birthdate - NEW: birthdate is now required
        bdate = data.get('bdate')
        if not bdate:
            print("Birthdate is missing")
            # ADD FAILED LOG:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Account Failed', 'Birthdate is required', current_user.account_id)
            return jsonify({
                "success": False, 
                "message": "Birthdate is required"
            }), 400

        user = data.get('username')
        password = data.get('password')  # This is now auto-generated from frontend
        email = data.get('email')

        fname = data.get('fname').upper()
        mname = data.get('mname', '').upper()
        lname = data.get('lname').upper()
        suffix = data.get('suffix', 'NA')

        contact = data.get('contact_no')
        acct_created = date.today().strftime("%Y-%m-%d")
        branch = data.get('branch')

        print(f"Processing account creation for: {fname} {lname} ({user})")
        print(f"Auto-generated password: {password}")

        # Check if username already exists
        print(f"Checking if username exists: {user}")
        if username_exists(user):
            print(f"Username already exists: {user}")
            # ADD FAILED LOG:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Account Failed', f'Username already exists: {user}', current_user.account_id)
            return jsonify({
                "success": False, 
                "message": "Username already exists. Please choose a different username."
            }), 400

        # Check if email already exists (if email provided)
        if email:
            print(f"Checking if email exists: {email}")
            if email_exists(email):
                print(f"Email already exists: {email}")
                # ADD FAILED LOG:
                db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Account Failed', f'Email already exists: {email}', current_user.account_id)
                return jsonify({
                    "success": False, 
                    "message": "Email address already exists. Please use a different email."
                }), 400

        # Create account using db_conn
        print("Calling db_conn.create_account...")
        create_new_acc = db_conn.create_account(
            user=user, 
            password=password, 
            email=email, 
            fname=fname, 
            mname=mname, 
            lname=lname,
            suffix=suffix, 
            bdate=bdate, 
            contact=contact, 
            acct_created=acct_created, 
            branch=branch
        )
        
        print(f"db_conn.create_account returned: {create_new_acc}")
    
        if create_new_acc is True:
            print("Account creation successful!")
            # ADD SUCCESS LOG:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Account', f'Created account for {fname} {lname} (Username: {user})', current_user.account_id)
            return jsonify({
                "success": True,
                "message": "Account created successfully!"
            })
        else:
            print(f"Account creation failed with: {create_new_acc}")
            # ADD FAILED LOG:
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Account Failed', f'Database error: {create_new_acc}', current_user.account_id)
            return jsonify({
                "success": False, 
                "message": f"Account creation failed: {create_new_acc}"
            }), 500

    except Exception as e:
        print(f"Error creating account: {str(e)}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        # ADD FAILED LOG FOR EXCEPTION:
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Create Account Error', f'Exception: {str(e)}', current_user.account_id)
        return jsonify({
            "success": False, 
            "message": "An unexpected error occurred. Please try again."
        }), 500
    
@server.route('/api/accounts/archive', methods=['PATCH'])
@login_required
@roles_required('admin', 'superadmin')
def archive_account():
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        print(f"🔍 ARCHIVE REQUEST - IDs: {ids}")
        
        if not ids:
            return jsonify({"success": False, "error": "No account IDs provided"}), 400

        # Use SINGLE session for ALL operations
        db_session = SessionLocal()
        success_count = 0
        failed_ids = []
        
        try:
            # BULK UPDATE - archive all accounts in one query
            result = db_session.query(models.Accounts).filter(
                models.Accounts.account_id.in_(ids)
            ).update({
                models.Accounts.acct_status: 'archived'
            }, synchronize_session=False)
            
            db_session.commit()
            success_count = result
            
            print(f"✅ Successfully archived {success_count} accounts in one operation")
            db_conn.POST_action_log(current_user.username, current_user.user_level, 'Archive Accounts', f'Archived {success_count} account(s)', current_user.account_id)
            
            # Log the bulk action
            if success_count > 0:
                try:
                    archiver = db_session.query(models.Accounts).filter(
                        models.Accounts.account_id == current_user.account_id
                    ).first()
                    
                    if archiver:
                        db_conn.POST_action_log(
                            archiver.username,
                            archiver.user_level,
                            "Bulk Archive Accounts",
                            f"Archived {success_count} account(s): {ids}",
                            current_user.account_id
                        )
                except Exception as log_error:
                    print(f"⚠️ Logging failed but archive succeeded: {log_error}")
            
        except Exception as e:
            db_session.rollback()
            raise e
        finally:
            db_session.close()
        
        print(f"📊 Archive summary: {success_count}/{len(ids)} successful")
        
        if success_count > 0:
            return jsonify({"success": True, "archived_count": success_count})
        else:
            return jsonify({"success": False, "error": "Failed to archive any accounts"}), 500
            
    except Exception as e:
        print(f"❌ Error in archive_account: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
# ==================== API ROUTES FOR AUDIT TRAILS ====================
# SIMPLE WORKING API ROUTES
@server.route('/api/get_users', methods=['GET'])
@login_required
@roles_required('admin', 'superadmin')
def api_get_users():
    try:
        print("🔍 Getting users...")
        db_session = db_conn.SessionLocal()
        users = db_session.query(db_conn.models.Accounts).filter(
            db_conn.models.Accounts.acct_status.in_(['approved', 'staff'])
        ).all()
        
        users_list = []
        for user in users:
            users_list.append({
                'user_id': user.account_id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name
            })
        
        print(f"✅ Found {len(users_list)} users")
        return jsonify(users_list)
        
    except Exception as e:
        print(f"❌ Error getting users: {e}")
        return jsonify([])
    finally:
        db_session.close()

@server.route('/api/get_filtered_logs', methods=['GET'])
@login_required
@roles_required('admin', 'superadmin')
def api_get_filtered_logs():
    try:
        user_id = request.args.get('user_id', '')
        filter_date = request.args.get('date', '')
        
        print(f"🔍 Filtering - User: {user_id}, Date: {filter_date}")
        
        # Get all logs
        all_logs = db_conn.GET_audit_logs()
        print(f"📋 Total logs: {len(all_logs)}")
        
        filtered_logs = []
        
        for log in all_logs:
            # Get user ID - try different attribute names
            log_user_id = getattr(log, 'account_id', None)
            if log_user_id is None:
                log_user_id = getattr(log, 'user_id', None)
            
            log_timestamp = getattr(log, 'timestamp', None)
            
            # Check user filter
            user_match = not user_id or (log_user_id and str(log_user_id) == user_id)
            
            # Check date filter
            date_match = True
            if filter_date:
                if log_timestamp:
                    if hasattr(log_timestamp, 'strftime'):
                        log_date = log_timestamp.strftime('%Y-%m-%d')
                    else:
                        log_date = str(log_timestamp)[:10]
                    date_match = (log_date == filter_date)
                else:
                    date_match = False
            
            if user_match and date_match:
                # Format timestamp
                if log_timestamp and hasattr(log_timestamp, 'strftime'):
                    timestamp_str = log_timestamp.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    timestamp_str = str(log_timestamp)
                
                filtered_logs.append({
                    'log_id': getattr(log, 'log_id', ''),
                    'username': getattr(log, 'username', ''),
                    'user_level': getattr(log, 'user_level', ''),
                    'action': getattr(log, 'action', ''),
                    'details': getattr(log, 'details', ''),
                    'timestamp': timestamp_str,
                    'ip_address': getattr(log, 'ip_address', '')
                })
        
        print(f"✅ Filtered logs: {len(filtered_logs)}")
        return jsonify({
            "success": True, 
            "logs": filtered_logs,
            "filters": {"user_id": user_id, "date": filter_date}
        })
        
    except Exception as e:
        print(f"❌ Error in get_filtered_logs: {e}")
        return jsonify({
            "success": False, 
            "error": str(e),
            "logs": []
        }), 500
    
    # DEBUG ROUTE - Check if APIs are working
@server.route('/api/debug_test')
def debug_test():
    return jsonify({"message": "API is working!", "status": "success"})

# DEBUG ROUTE - Check audit logs structure
@server.route('/api/debug_logs_info')
@login_required
@roles_required('admin', 'superadmin')
def debug_logs_info():
    try:
        logs = db_conn.GET_audit_logs()
        if not logs:
            return jsonify({"message": "No logs found", "count": 0})
        
        # Check first log structure
        first_log = logs[0]
        log_attrs = {}
        for attr in dir(first_log):
            if not attr.startswith('_'):
                try:
                    value = getattr(first_log, attr)
                    log_attrs[attr] = {
                        "type": str(type(value)),
                        "value": str(value)[:100] if value else "None"
                    }
                except:
                    pass
        
        return jsonify({
            "total_logs": len(logs),
            "first_log_attributes": log_attrs,
            "sample_data": {
                "log_id": getattr(first_log, 'log_id', 'N/A'),
                "username": getattr(first_log, 'username', 'N/A'),
                "account_id": getattr(first_log, 'account_id', 'N/A'),
                "timestamp": str(getattr(first_log, 'timestamp', 'N/A'))
            }
        })
        
    except Exception as e:
        return jsonify({"error": str(e)})
    
# =============================================
# TARGET VS ACTUAL API ROUTES
# =============================================

@server.route('/api/get_targets/<int:year>')
@login_required
@roles_required('admin', 'superadmin')
def get_targets(year):
    """Get targets for a specific year"""
    db_session = SessionLocal()
    try:
        # Query your targets table - adjust the model name as needed
        targets_row = db_session.query(models.Report_TvA).filter(models.Report_TvA.year == year).first()
        
        if targets_row:
            return jsonify({
                "Classic": targets_row.classic or 0,
                "Bronze": targets_row.bronze or 0,
                "Silver": targets_row.silver or 0,
                "Gold": targets_row.gold or 0,
                "Platinum": targets_row.platinum or 0,
                "Safe Card": targets_row.safe_card or 0,
                "Senior": targets_row.senior or 0,
                "Senior+": targets_row.senior_plus or 0
            })
        else:
            # Return zeros if no targets set
            return jsonify({
                "Classic": 0, "Bronze": 0, "Silver": 0, "Gold": 0,
                "Platinum": 0, "Safe Card": 0, "Senior": 0, "Senior+": 0
            })
    except Exception as e:
        print(f"Error getting targets: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        db_session.close()

@server.route('/api/save_targets', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def save_targets():
    """Save targets for a specific year"""
    data = request.get_json()
    year = data.get('year')
    targets = data.get('targets', {})
    
    print(f"💾 Saving targets for year {year}: {targets}")
    
    db_session = SessionLocal()
    try:
        # Check if targets exist for this year
        existing = db_session.query(models.Report_TvA).filter(models.Report_TvA.year == year).first()
        
        if existing:
            # Update existing
            existing.classic = targets.get('Classic', 0)
            existing.bronze = targets.get('Bronze', 0)
            existing.silver = targets.get('Silver', 0)
            existing.gold = targets.get('Gold', 0)
            existing.platinum = targets.get('Platinum', 0)
            existing.safe_card = targets.get('Safe Card', 0)
            existing.senior = targets.get('Senior', 0)
            existing.senior_plus = targets.get('Senior+', 0)
            print(f"✅ Updated existing targets for year {year}")
        else:
            # Create new
            new_targets = models.Report_TvA(
                year=year,
                classic=targets.get('Classic', 0),
                bronze=targets.get('Bronze', 0),
                silver=targets.get('Silver', 0),
                gold=targets.get('Gold', 0),
                platinum=targets.get('Platinum', 0),
                safe_card=targets.get('Safe Card', 0),
                senior=targets.get('Senior', 0),
                senior_plus=targets.get('Senior+', 0)
            )
            db_session.add(new_targets)
            print(f"✅ Created new targets for year {year}")
        
        db_session.commit()
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Save Targets', f'Saved targets for year {year}', current_user.account_id)
        return jsonify({"success": True})
    except Exception as e:
        db_session.rollback()
        print(f"❌ Error saving targets: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        db_session.close()

# =============================================
# BUDGET VS EXPENSES API ROUTES
# =============================================

@server.route('/api/get_budget/<int:year>')
@login_required
@roles_required('admin', 'superadmin')
def get_budget(year):
    """Get budget data for a specific year"""
    try:
        # For now, return sample data - you can replace this with actual database queries
        sample_budget_data = [
            {
                "id": 1,
                "account_code": "5020401",
                "account_name": "Gasoline & Oil",
                "budget": 100800,
                "jan": 1400, "feb": 4100, "mar": 3400, "apr": 3700, "may": 2000, "jun": 2100,
                "jul": 1400, "aug": 0, "sep": 0, "oct": 0, "nov": 0, "dec": 0,
                "total_expense": 18100,
                "balance": 82700,
            },
            {
                "id": 2,
                "account_code": "5020511",
                "account_name": "Travel Local",
                "budget": 20000,
                "jan": 0, "feb": 0, "mar": 0, "apr": 0, "may": 60, "jun": 20,
                "jul": 0, "aug": 0, "sep": 0, "oct": 0, "nov": 0, "dec": 0,
                "total_expense": 80,
                "balance": 19920,
            },
            {
                "id": 3,
                "account_code": "5020601",
                "account_name": "Meals and Snacks",
                "budget": 20000,
                "jan": 405, "feb": 1905, "mar": 1520, "apr": 2209, "may": 1639, "jun": 1372,
                "jul": 400, "aug": 0, "sep": 0, "oct": 0, "nov": 0, "dec": 0,
                "total_expense": 9450,
                "balance": 10550,
            }
        ]
        return jsonify(sample_budget_data)
    except Exception as e:
        print(f"Error getting budget: {e}")
        return jsonify({"error": str(e)}), 500

@server.route('/api/save_budget', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def save_budget():
    """Save budget data"""
    data = request.get_json()
    year = data.get('year')
    budget_data = data.get('budget', [])
    
    print(f"💾 Saving budget for year {year}")
    print(f"📊 Budget data: {budget_data}")
    
    try:
        # For now, just log the data - you can implement actual database saving later
        db_conn.POST_action_log(current_user.username, current_user.user_level, 'Save Budget', f'Saved budget for year {year} with {len(budget_data)} items', current_user.account_id)
        return jsonify({"success": True, "message": "Budget data received (not yet saved to database)"})
    except Exception as e:
        print(f"❌ Error saving budget: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    
#? -------------------- NOTIFICATIONS -------------------- ?#

@server.route('/api/notifications', methods=['GET'])
def get_notifications():
    try:
        from datetime import datetime, timedelta
        today = datetime.now().date()
        notifications = []
        
        print("🔍 Starting notifications check...")
        
        # Get all member records
        all_records = db_conn.get_member_records()
        print(f"📊 Found {len(all_records)} records")
        
        # 1. Check for birthdays today
        birthday_count = 0
        for record in all_records:
            try:
                # Get record ID safely
                record_id = getattr(record, 'record_id', getattr(record, 'id', None))
                if not record_id:
                    continue
                    
                # Get entries for this record
                entries = db_conn.get_entries(record_id)
                
                for entry in entries:
                    # Check if entry has birth_date
                    birth_date = getattr(entry, 'birth_date', None)
                    if birth_date:
                        # Convert to date object if string
                        if isinstance(birth_date, str):
                            try:
                                birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()
                            except:
                                continue
                        elif isinstance(birth_date, datetime):
                            birth_date = birth_date.date()
                        
                        # Check if birthday is today
                        if birth_date.month == today.month and birth_date.day == today.day:
                            first_name = getattr(entry, 'first_name', '')
                            last_name = getattr(entry, 'last_name', '')
                            email = getattr(entry, 'email', '')
                            
                            notifications.append({
                                'type': 'birthday',
                                'message': f"🎂 {first_name} {last_name} has birthday today!",
                                'member_name': f"{first_name} {last_name}",
                                'member_email': email,
                                'priority': 'high'
                            })
                            birthday_count += 1
            except Exception as e:
                print(f"❌ Error processing record: {e}")
                continue
        
        print(f"🎂 Found {birthday_count} birthdays today")
        
        # 2. Check for expiring memberships (30 days)
        expiring_count = 0
        thirty_days_from_now = today + timedelta(days=30)
        
        for record in all_records:
            try:
                record_id = getattr(record, 'record_id', getattr(record, 'id', None))
                if not record_id:
                    continue
                    
                entries = db_conn.get_entries(record_id)
                
                for entry in entries:
                    or_date = getattr(entry, 'OR_date', None)
                    if or_date:
                        # Convert to date object if string
                        if isinstance(or_date, str):
                            try:
                                or_date = datetime.strptime(or_date, '%Y-%m-%d').date()
                            except:
                                continue
                        elif isinstance(or_date, datetime):
                            or_date = or_date.date()
                        
                        # Calculate expiration (OR_date + 1 year)
                        expiration_date = or_date + timedelta(days=365)
                        days_until_expiry = (expiration_date - today).days
                        
                        # Check if expiring within 30 days
                        if 0 <= days_until_expiry <= 30:
                            first_name = getattr(entry, 'first_name', '')
                            last_name = getattr(entry, 'last_name', '')
                            email = getattr(entry, 'email', '')
                            
                            notifications.append({
                                'type': 'expiring',
                                'message': f"⏰ {first_name} {last_name} membership expires in {days_until_expiry} days",
                                'member_name': f"{first_name} {last_name}",
                                'member_email': email,
                                'additional_data': {'days_left': days_until_expiry},
                                'priority': 'medium'
                            })
                            expiring_count += 1
            except Exception as e:
                print(f"❌ Error processing record for expiring: {e}")
                continue
        
        print(f"⏰ Found {expiring_count} expiring memberships")
        print(f"📨 Total notifications: {len(notifications)}")
        
        return jsonify({
            'success': True,
            'notifications': notifications,
            'total_count': len(notifications),
            'summary': {
                'birthdays': birthday_count,
                'expiring': expiring_count
            }
        })
        
    except Exception as e:
        print(f"❌ Major error in get_notifications: {e}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'error': str(e),
            'notifications': [],
            'total_count': 0
        }), 500

# TEST ROUTE - OPTIONAL
@server.route('/api/notifications/test', methods=['GET'])
def test_notifications_route():
    return jsonify({"message": "Notifications route is working", "success": True})

#? -------------------- MISC ROUTES -------------------- ?#

@server.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404

@server.errorhandler(403)
def page_not_found(e):
    return render_template("403.html"), 403

ROOT_STATIC_FILES = {
    "robots.txt",
    "humans.txt",
    "security.txt"
}
@server.route('/<path:filename>')
def root_static_files(filename):
    if filename in ROOT_STATIC_FILES:
        return send_from_directory(server.static_folder, filename)
    abort(404)
    
# Favicon
@server.route('/favicon.ico')
def favicon():
    return send_from_directory(server.static_folder, 'assets/favicon.ico') 

#? -------------------- END -------------------- ?#


if __name__ == '__main__':
    
    # DEPRECATED LIVE RELOAD METHOD. REMOVE LATER
    '''
    flask_server = Server(server.wsgi_app)
    flask_server.watch('static/*.*')  # watches static files (CSS/JS)
    flask_server.watch('templates/*.html')  # watches templates
    flask_server.serve(port=5000, host="127.0.0.1")

    #server.run(debug=True, use_reloader=True, port=5000)
    '''
    print("FLASK_ENV:", os.getenv("FLASK_ENV"))
    
    if os.getenv("FLASK_ENV") == "production":
        #server.run(host="0.0.0.0", port=5000)
        pass
    else:
    #if os.getenv("FLASK_ENV") == "development":
        flask_server = Server(server.wsgi_app)
        flask_server.watch('static/*.*')
        flask_server.watch('templates/*.html')
        flask_server.serve(port=5000, host="127.0.0.1")